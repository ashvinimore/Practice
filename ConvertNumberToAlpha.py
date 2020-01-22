import mysql.connector
from openpyxl import load_workbook


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string





#output:AB
try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='Zeus',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
        file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
        wb = load_workbook('/home/pavan/Music/Data Warehouse/DWH-1.xlsx')
        # print(wb.get_sheet_names())
        # anotherSheet = wb.active
        #
        # match_fetch = {}
        sheet = wb.get_sheet_by_name('Industry Template')
        max_col = (sheet.max_column)
        # print(max_col)
        max_row = (sheet.max_row)
except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")

def mainf(max_col):
    rowconst = ''
    flag = 0
    for col in range(1, max_col):
        letter = (colnum_string(col))
        # print(letter)
        for row in range(1, max_row):
            row_ex = ''
            row_ex = letter + str(row)
            if(sheet[row_ex].value == 'GXBRL ID'):
                rowconst = row
                rowval = letter + str(rowconst)
                flag = 1
                break
        if flag == 1:
            # print(rowconst)
            break
    rowseg = ''
    roegic = ''
    rowgeo = ''

    for col in range(1, max_col):
        letter = (colnum_string(col))
        row_ex = letter + str(rowconst)
        if(sheet[row_ex].value  == 'SegmentID'):
            rowseg = letter + str(rowconst)
        elif (sheet[row_ex].value  == 'GICSID'):
            roegic = letter + str(rowconst)
        elif (sheet[row_ex].value == 'Geography'):
            rowgeo = letter + str(rowconst)
        elif(sheet[row_ex].value == 'ProductID'):
            rowprod =  letter + str(rowconst)
    gxbrl = sheet[rowval].value
    segments = sheet[rowseg].value
    gicid = sheet[roegic].value
    prodid = sheet[rowprod].value
    geoid = sheet[rowgeo].value
    # print(gxbrl,segments,gicid,prodid,geoid)
    autoid = 1
    if (gxbrl != None or segments != None or gicid != None or geoid != None):
        if (gxbrl == None):
            gxbrl = None
            gxbrl = str(gxbrl)
        if (segments == None):
            segments = None
            segments = str(segments)
        if (gicid == None):
            gicid = None
            gicid = str(gicid)
        if (geoid == None):
            geoid = None
            geoid = str(geoid)
        if (prodid == None):
            prodid = None
            prodid = str(prodid)

        cursor.execute(
            "SELECT * FROM  DWHMainQuery1 WHERE  GXbrlEleID = %s and SegmentID = %s and GeoID = %s and GicID = %s and ProductID = %s",
            (gxbrl, segments, geoid, gicid, prodid))
        result = cursor.fetchall()

        if (result == []):
            parent = None
            autoid = autoid + 1
            UID = autoid
            print(UID, gxbrl, segments, geoid, gicid, prodid)
            insertQuery = "insert into DWHMainQuery1(UID,GXbrlEleID,SegmentID,GeoID,GicID,ProductID)  values(" \
                          "'%s','%s','%s','%s','%s','%s')" % (
                              (UID, gxbrl, segments, geoid, gicid, prodid)
                          )
            # insertQuery = "insert into GlobalElement(GXbrlEleID,Description) values('1','1')"
            try:
                cursor.execute(insertQuery)
                print("query excuted successfully")
                connection.commit()
            except Exception as e:
                print(e)

        #
        #
        else:
            print("already ")

mainf(max_col)
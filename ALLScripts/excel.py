
from openpyxl import load_workbook
from openpyxl.styles import  Alignment, Protection, Font
import csv
import mysql.connector
from mysql.connector import Error
#To load workbook
wb = load_workbook('/home/pavan/Music/Data Warehouse/GE/GlobalElement_LiveExported_190820191.xlsx')
# print(wb.get_sheet_names())

try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='DataWarehouse',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
        cursor.execute("select * from period")

        try:
            sheet = wb.get_sheet_by_name('GlobalElement_LiveExported_19082019')
            print(sheet.max_column,sheet.max_row)
            max_row = sheet.max_row
            max_col =  sheet.max_column
            try:
                # print("in try")
                # record1 = cursor.fetchone()
                # print("Your connected to - ", record1)
                for row in range(2,max_row+1):
                    for col in range(1,max_col+1):
                        print(col[0])
                        GXbrlEleID  = col[0]
                        Description = col[1]
                        Taxonomy = col[2]
                        FormID = col[3]
                        DataID = col[4]
                        Unit = col[5]
                        ParentGXbrlEleID = col[6]
                        Weight  = col[7]
                        Formula = col[8]

                        insertQuery = "insert into GlobalElement(GXbrlEleID,Description,Taxonomy,FormID,DataID,Unit,ParentGXbrlEleID,Weight,Formula)  values(" \
                                                                  "%s,%s,'%s,%s,%s,%s,%s,%s, %s)" % (
                                                                GXbrlEleID, Description, Taxonomy, FormID, DataID, Unit, ParentGXbrlEleID, Weight, Formula
                                                              )
                        # insertQuery = "insert into GlobalElement(GXbrlEleID,Description) values('1','1')"
                        try:
                            cursor.execute(insertQuery)
                            print("query excuted successfully")
                            connection.commit()
                        except Exception as e:
                            print(e)
                        break
            #

            except Exception as e:
                print(e)
        except Exception as e:
            print(e)

except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")
#


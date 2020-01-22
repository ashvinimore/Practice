import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook
from openpyxl.styles import  Alignment, Protection, Font

#To load workbook
wb = load_workbook('./AugustReport/July2019IEAReport.xlsx')
print(wb.get_sheet_names())
anotherSheet = wb.active

# try:
#     sheet = wb.get_sheet_by_name('Annual_Previous')
#     wb.remove_sheet(sheet)
#     sheet = wb.get_sheet_by_name('Annual')
#     ws2 = wb.copy_worksheet(sheet)
#     ws2.title = 'Annual_Previous'
# except Exception as e:
#     sheet = wb.get_sheet_by_name('Annual')
#     ws2 = wb.copy_worksheet(sheet)
#     ws2.title = 'Annual_Previous'

# try:
#     sheet = wb.get_sheet_by_name('Quarter_Previous')
#     wb.remove_sheet(sheet)
#     sheet = wb.get_sheet_by_name('Quarter')
#     ws2 = wb.copy_worksheet(sheet)
#     ws2.title = 'Quarter_Previous'
# except Exception as e:
#     sheet = wb.get_sheet_by_name('Quarter')
#     ws2 = wb.copy_worksheet(sheet)
#     ws2.title = "Quarter_Previous"
#     pass

wb.save('./July2019IEAReport.xlsx')

#to find row number of OECD,NONOOECD,TOTALOECD,TOTALNONOECD
def fetch_excel(sheet):
    print(sheet)
    flag = 0
    mbd = -1
    fetch = {}
    match_fetch={}
    sheet = wb.get_sheet_by_name(sheet)
    max_col = (sheet.max_column)
    print(max_col)
    max_row = (sheet.max_row)
    print(max_row)
    for row in range(1,max_row):
        row_ex = 'A'+str(row)
        if sheet[row_ex].value == 'Demand':
            fetch['Demand'] = row
            match_fetch['Demand'] = row
            row = row + 1
            row_ex = 'A' + str(row)
            if sheet[row_ex].value == 'OECD':
                match_fetch['DemandOECD'] = row
                row = row + 1
                oecd_demand = row
                fetch['oecd_demand'] = oecd_demand

            for total in range(row,max_row):
                row_ex = 'A' + str(total)
                if sheet[row_ex].value == 'Total OECD':
                    totaloecd_demand = total
                    total = total+1
                    print('totaloecd_demand',totaloecd_demand)
                    fetch['totaloecd_demand'] = totaloecd_demand
                    match_fetch['DemandTotal OECD'] = totaloecd_demand
                    break
            for non in range(total,max_row):
                row_ex = 'A' + str(non)
                if sheet[row_ex].value == 'Non-OECD':
                    match_fetch['DemandNon-OECD'] = non
                    non = non + 1
                    nonoecd_demand = non
                    fetch['nonoecd_demand'] = nonoecd_demand

                    break
            for total_non in range(non,max_row):
                row_ex = 'A' + str(total_non)
                if sheet[row_ex].value == 'Total Non-OECD':
                    totalnonoecd_demand = total_non
                    total_non = total_non + 1
                    fetch['totalnonoecd_demand'] = totalnonoecd_demand
                    match_fetch['DemandTotal Non-OECD'] = totalnonoecd_demand
                    break
        elif sheet[row_ex].value == 'Supply':
            fetch['Supply'] = row
            match_fetch['Supply'] = row
            row = row + 1
            row_ex = 'A' + str(row)
            if sheet[row_ex].value == 'OECD':               
                row = row + 1
                oecd_supply = row
                fetch['oecd_supply'] = oecd_supply
            for total in range(row,max_row):
                row_ex = 'A' + str(total)
                if sheet[row_ex].value == 'Total OECD':
                    totaloecd_supply = total
                    total = total+1
                    fetch['totaloecd_supply'] = totaloecd_supply
                    match_fetch['SupplyTotal OECD'] = totaloecd_supply
                    break
            for non in range(total,max_row):
                row_ex = 'A' + str(non)
                if sheet[row_ex].value == 'Non-OECD':
                    match_fetch['SupplyNon-OECD'] = non
                    nonoecd_supply = non
                    non = non + 1
                    fetch['nonoecd_supply'] = nonoecd_supply

                    break
            for total_non in range(non,max_row):
                row_ex = 'A' + str(total_non)
                if sheet[row_ex].value == 'Total Non-OECD':
                    totalnonoecd_supply = total_non
                    total_non = total_non + 1
                    fetch['totalnonoecd_supply'] = totalnonoecd_supply
                    match_fetch['SupplyTotal Non-OECD'] = totalnonoecd_supply
                    break
            for tot_n_opec in range(non, max_row):
                row_ex = 'A' + str(tot_n_opec)
                if sheet[row_ex].value == 'Total Non-OPEC Supply':
                    total_non_opec_supply = tot_n_opec
                    tot_n_opec = tot_n_opec + 1
                    fetch['total_non_opec_supply'] = total_non_opec_supply
                    match_fetch['Total Non-OPEC Supply'] = total_non_opec_supply
                    break
            for opec in range(tot_n_opec, max_row):
                row_ex = 'A' + str(opec)
                if sheet[row_ex].value == 'OPEC':
                    match_fetch['OPEC'] = opec
                    opec_supply = opec
                    fetch['opec_supply'] = opec_supply
                    opec = opec + 1
                    break
            for tot_opec in range(opec, max_row):
                row_ex = 'A' + str(tot_opec)
                if sheet[row_ex].value == 'Total OPEC':
                    total_opec_supply = tot_opec
                    fetch['total_opec_supply'] = total_opec_supply
                    match_fetch['Total OPEC'] = total_opec_supply
                    tot_opec = tot_opec + 1
                    break
        elif sheet[row_ex].value == 'Stocks':
            row = row + 1
            row_ex = 'A' + str(row)
            if sheet[row_ex].value == 'Government Controlled':
                stock_govt = row
                print('stock_gov',stock_govt)

                for sto_oecd in range(stock_govt,max_row):
                    row_ex = 'A' + str(sto_oecd)
                    if sheet[row_ex].value == 'OECD':
                        oecd_stock = sto_oecd+1
                        fetch['oecd_stock'] = oecd_stock
                        break
        elif sheet[row_ex].value == 'Industry Stocks':
            idstock = row
            row = row + 1

            for ioecd in range(row, max_row):
                row_ex = 'A' + str(ioecd)
                if sheet[row_ex].value == 'OECD':
                    ioecd_stock = ioecd+1
                    fetch['ioecd_stock'] = ioecd_stock
                    print('ioecd_stock', ioecd_stock)

                if (sheet[row_ex].value == 'CHANGE - MBD'):
                    mbd = ioecd
                    print('CHANGE - MBD', row_ex)
                    break
    #important condition to limit the row
        if (mbd == row):
            break

    return fetch,match_fetch

#
# #
# #
# #Database Connection
try:
    connection = mysql.connector.connect(host='13.229.23.103',
                                         database='Aletheia',
                                         user='root',
                                         password='AP@2019capio!')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)

    def callproc_script():
        try:
            cursor.callproc('DP_IEAReportDataPopulation')
            for result in cursor.stored_results():
                print(result.fetchone())
        except Error as e:
            print("print", e)

#    #dynamically access Regions and Countries
    def regions(sheet):
        print(sheet)
        # wb.create_sheet(sheet)
        sheet = wb.get_sheet_by_name(sheet)
        max_col = (sheet.max_column)
        max_row = (sheet.max_row) + 1
        print(max_col)  #
        #     # print(max_row)
        flag = 1
        nonoced = 0
        # font = Font(name='Calibri',size = 11,bold = True)

        gr_demand = ['OECD', 'Non-OECD']
        gr_supply = ['Non-OPEC-OECD', 'Non-OPEC-Non-OECD']
        try:
            for groups in gr_demand:
                if flag == 1:
                    rows = 10
                    sheet.cell(row=9, column=1).value = 'Demand'
                    for row in range(rows, max_row):
                        row_ex = 'A' + str(row)
                        if sheet[row_ex].value == 'Demand':
                            row = row + 1
                            sheet.cell(row=row, column=1).value = groups
                            sheet.cell(row=row, column=1).font = Font(bold=True)
                            rows = row + 1
                            break


                elif flag == 2:
                    rows = nonoced
                    sheet.cell(row=rows, column=1).value = groups
                    sheet.cell(row=rows, column=1).font = Font(bold=True)
                    rows = rows + 1

                else:
                    pass

                cursor.execute(
                    "SELECT distinct RegionName FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = %s",
                    (groups,))
                result = cursor.fetchall()
                row = row + 1
                for value in result:
                    row = row + 1
                    # print(row)
                    val = value[0]
                    sheet.cell(row=row, column=1).value = val
                    sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
                    sheet.cell(row=row, column=1).font = Font(bold=True)
                    cursor.execute(
                        "select distinct Country from ReportFor2b where RegionName = %s and Gropus = %s", (val,groups,))
                    result_country = cursor.fetchall()
                    for value in result_country:
                        row = row + 1
                        print(value[0])
                        sheet.cell(row=row, column=1).value = value[0]
                        sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=2, wrap_text=True)
                        sheet.cell(row=row, column=1).font = Font(bold=False)
                sheet.cell(row=row, column=1).value = 'Total' + groups
                sheet.cell(row=row, column=1).font = Font(bold=True)
                sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=0, wrap_text=True)
                nonoced = row + 1
                flag = 2

            print(row)
            row = row + 1
            sheet.cell(row=row, column=1).value = ' '
            row = row + 1
            sheet.cell(row=row, column=1).value = 'Supply'
            sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=0, wrap_text=True)
            sheet.cell(row=row, column=1).font = Font(bold=True)
            flag = 1

            for groups in gr_supply:
                if flag == 1:
                    rows = row
                    for row in range(rows, max_row):
                        row_ex = 'A' + str(row)
                        if sheet[row_ex].value == 'Supply':
                            row = row + 1
                            sheet.cell(row=row, column=1).value = groups
                            sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=0,
                                                                                wrap_text=True)
                            sheet.cell(row=row, column=1).font = Font(bold=True)
                            rows = row + 1
                            break

                elif flag == 2:
                    rows = nonoced
                    sheet.cell(row=rows, column=1).value = groups
                    sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=0, wrap_text=True)
                    sheet.cell(row=row, column=1).font = Font(bold=True)
                else:
                    pass

                cursor.execute(
                    "SELECT distinct RegionName FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = %s",
                    (groups,))
                result = cursor.fetchall()
                for value in result:
                    row = row + 1
                    # print(row)
                    val = value[0]
                    sheet.cell(row=row, column=1).value = val
                    sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
                    sheet.cell(row=row, column=1).font = Font(bold=True)
                    cursor.execute(
                        "select distinct Country from ReportFor3 where RegionName = %s and Groups = %s ", (val,groups,))
                    result_country = cursor.fetchall()
                    for value in result_country:
                        row = row + 1
                        print(value[0])
                        sheet.cell(row=row, column=1).value = value[0]
                        sheet.cell(row=row, column=1).font = Font(bold=False)
                        sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=2, wrap_text=True)
                sheet.cell(row=row, column=1).value = 'Total' + groups.replace('Non-OPEC-', '')
                sheet.cell(row=row, column=1).font = Font(bold=True)
                nonoced = row
                flag = 2

            print(row)
            row = row + 1
            sheet.cell(row=row, column=1).value = 'Processing gains'
            row = row + 1
            sheet.cell(row=row, column=1).value = 'Global Biofuels'
            row = row + 1
            sheet.cell(row=row, column=1).value = 'Total Non-OPEC Supply'
            sheet.cell(row=row, column=1).font = Font(bold=True)
            row = row + 1
            sheet.cell(row=row, column=1).value = 'Non-OPEC: Historical Composition'
            sheet.cell(row=row, column=1).font = Font(bold=True)
            row = row + 1
            sheet.cell(row=row, column=1).value = ''
            row = row + 1
            sheet.cell(row=row, column=1).value = 'OPEC'
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=0, wrap_text=True)

            if sheet.cell(row=row, column=1).value == 'OPEC':
                print("OPECCCCCCCCCCCCCCCC")
                cursor.execute(
                    "SELECT distinct Country FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'OPEC'")

                result = cursor.fetchall()
                for value in result:
                    print(value[0])
                    row = row + 1
                    sheet.cell(row=row, column=1).value = value[0]
                    sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=2, wrap_text=True)

            sheet.cell(row=row, column=1).value = 'Crude'
            sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
            row = row + 1
            sheet.cell(row=row, column=1).value = 'NGL'
            row = row + 1
            sheet.cell(row=row, column=1).value = 'Total OPEC'
            row = row + 1
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=1).value = 'OPEC: Historical Composition'
            sheet.cell(row=row, column=1).font = Font(bold=True)
            row = row + 1
            sheet.cell(row=row, column=1).value = 'Total Supply'
            sheet.cell(row=row, column=1).font = Font(bold=True)
            row = row + 1
        except Error as e:
            print("print", e)
        wb.save('./IEA QUARTERLY FILE_WITH MODEL_PW_Country.xlsx')

    def quarter_data():
        # try:
        #     sheet = wb.get_sheet_by_name('Quarter_Previous')
        #     wb.remove_sheet(sheet)
        #     sheet = wb.get_sheet_by_name('Quarter')
        #     ws2 = wb.copy_worksheet(sheet)
        #     ws2.title = 'Quarter_Previous'
        #     wb.remove_sheet(sheet)
        # except Exception as e:
        #     sheet = wb.get_sheet_by_name('Quarter')
        #     ws2 = wb.copy_worksheet(sheet)
        #     ws2.title = 'Quarter_Previous'
        #     wb.remove_sheet(sheet)
        #
        # region = regions('Quarter')
        fetch = fetch_excel('Quarter')
        fetch = fetch[0]
        sheet = wb.get_sheet_by_name('Quarter')
        max_col = (sheet.max_column)
        print(max_col)
        max_row = (sheet.max_row)

        # reporteddate = '00:00:00
        try:
            #
#             for col in range(2, max_col):
#                     period = sheet.cell(row=1, column=col).value
#                     print(period)
#                     if (period == '2025-4Q'):
#                         break
#                     totaloecd_demand = int(fetch.get('totaloecd_demand')) + 1
#                     oecd_demand = int(fetch.get('oecd_demand'))
#                     for row in range(oecd_demand, totaloecd_demand):
#                         row_ex = 'A' + str(row)
#                         if (period == '2025-4Q'):
#                             flag = 1
#                             break
#                         else:
#                             region = sheet[row_ex].value
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'OECD' and Period = %s and Regionname = %s and ReportedDate like  '2019-02-13%'",
#                                 (period, region,))
#                             result = cursor.fetchall()
#                             if result and len(result) >0:
#                                 val = result[0][0]
#                                 if ('Q' in period):
#                                     sheet.cell(row=row, column=col).value = val
#                                     currentCell = sheet.cell(row=row, column=1)
#                                     # print(row)
#                                     currentCell.alignment = Alignment(horizontal='left', indent=1, vertical='bottom',
#                                                                       wrap_text=True)
#
#                             else:
#                                 country = sheet[row_ex].value
#                                 print(country)
#                                 cursor.execute(
#                                     "SELECT value FROM ReportFor2b WHERE ElementType = 'Demand' and Groups = 'OECD' and Period = %s and Country = %s and ReportedDate like  '2019-02-13%'",
#                                     (period, country,))
#                                 result = cursor.fetchall()
#                                 if result and len(result) > 0:
#                                     val = result[0][0]
#                                     if ('Q' in period):
#                                         sheet.cell(row=row, column=col).value = val
#                                         currentCell = sheet.cell(row=row, column=1)
#                                         # print(row)
#                                         currentCell.alignment = Alignment(horizontal='left', indent=2,
#                                                                           vertical='bottom',
#                                                                           wrap_text=True)
#
#                                 else:
#                                     pass
#                         if sheet[row_ex].value == 'Total OECD':
#                            # print("total oecds")
#                             currentCell = sheet.cell(row=row, column=1)
#                            # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                              vertical='bottom',
#                                                              wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'OECD' and Period = %s and ReportedDate like  '2019-02-13%' and Regionname IS NULL",
#                                 (period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=row, column=col).value = val
#                             else:
#                                 if ('Q' in period):
#                                     for val in result:
#                                         sheet.cell(row=row, column=col).value = val
# #
# # #
#             for col in range(2, max_col):
#                     period = sheet.cell(row=1, column=col).value
#                     print(period)
#                     if (period == '2025-4Q'):
#                         break
#     #             #for demand and nonoecd    #
#                     totalnonoecd_demand =  int(fetch.get('totalnonoecd_demand')) + 2
#                     #for total demand
#                     nonoecd_demand = int(fetch.get('nonoecd_demand'))
#                     for non in range(nonoecd_demand, totalnonoecd_demand):
#                         row_ex = 'A' + str(non)
#                         if sheet[row_ex].value == 'Total Non-OECD':
#                             currentCell = sheet.cell(row=non, column=1)
#                             # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#         #                         #print("total non oecds")
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'Non-OECD' and Period = %s and Regionname IS NULL and ReportedDate like  '2019-02-13%'",
#                                 (period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=non, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=non, column=col).value = val
#                         else:
#                             row_ex = 'A' + str(non)
#                             region = sheet[row_ex].value
#                             cursor.execute(
#                                 "SELECT Value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'Non-OECD' and Period = %s and RegionName = %s and ReportedDate like  '2019-02-13%'",
#                                 (period, region,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=non, column=col).value = val
#                                 currentCell = sheet.cell(row=non, column=1)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=1,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=non, column=col).value = val
#                                     currentCell = sheet.cell(row=non, column=1)
#                                     currentCell.alignment = Alignment(horizontal='left', indent=1,
#                                                                       vertical='bottom',
#                                                                       wrap_text=True)
#                         if sheet[row_ex].value == 'Total Demand':
#                             currentCell = sheet.cell(row=non, column=1)
#                             # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups IS NULL and Period = %s and Regionname IS NULL and ReportedDate like  '2019-02-13%'",
#                                 (period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=non, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=non, column=col).value = val
#                             break
#         #
#             for col in range(2, max_col):
#                     period = sheet.cell(row=1, column=col).value
#                     print(period)
#                     if (period == '2025-4Q'):
#                         break
#                     totaloecd_supply = int(fetch.get('totaloecd_supply')) +1
#                     oecd_supply =  int(fetch.get('oecd_supply'))
#                     for row in range(oecd_supply, totaloecd_supply):
#                         row_ex = 'A' + str(row)
#                         if sheet[row_ex].value == 'Total OECD':
#                             # print("total supply oecds")
#                             currentCell = sheet.cell(row=row, column=1)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-OECD' and Period = %s and Regionname IS NULL ",
#                                 (period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=row, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=row, column=col).value = val
#
#                             break
#                         else:
#                             row_ex = 'A' + str(row)
#                             region = sheet[row_ex].value
#                             currentCell = sheet.cell(row=row, column=1)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-OECD'  and Regionname = %s and Period = %s ",
#                                 (region,period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 country = sheet[row_ex].value
#                                 currentCell = sheet.cell(row=row, column=1)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=2,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT Value FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-OECD'  and Period = %s and Country = %s ",
#                                     (period, country,))
#                                 result = cursor.fetchall()
#                                 if result and len(result) > 0:
#                                     val = result[0][0]
#                                     sheet.cell(row=row, column=col).value = val
#
#                                 else:
#                                     country = sheet[row_ex].value  ##for uk
#                                     currentCell.alignment = Alignment(horizontal='left', indent=2,
#                                                                       vertical='bottom', wrap_text=True)
#                                     cursor.execute(
#                                         "SELECT Value FROM Report3 WHERE ElementType = 'Supply' and Period = %s and Country = %s",
#                                         (period, country,))
#                                     result = cursor.fetchall()
#                                     if result and len(result) > 0:
#                                         val = result[0][0]
#                                         sheet.cell(row=row, column=col).value = val
#
#                                     else:
#                                         pass
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=row, column=col).value = val
#                                     currentCell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
#                                     print(val, period,country)
#     # #
            for col in range(2, max_col):
                    period = sheet.cell(row=1, column=col).value
                    print(period)
                    if (period == '2025-1Q'):
                        break
                    else:# for supply non-oecd
                        totalnonoecd_supply = int(fetch.get('totalnonoecd_supply')) + 1
                        nonoecd_supply = int(fetch.get('nonoecd_supply'))
                        for non in range(nonoecd_supply, totalnonoecd_supply):
                            row_ex = 'A' + str(non)
                            if sheet[row_ex].value == 'Total Non-OECD':
                                currentCell = sheet.cell(row=non, column=1)
                                # print(row)
                                currentCell.alignment = Alignment(horizontal='left', indent=0,
                                                                  vertical='bottom',
                                                                  wrap_text=True)
                                cursor.execute(
                                    "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Period = %s and Regionname IS NULL",
                                    (period,))
                                result = cursor.fetchone()
                                if result is None:
                                    val = ''
                                    sheet.cell(row=non, column=col).value = val
                                else:
                                    for val in result:
                                        sheet.cell(row=non, column=col).value = val
                                break
                            else:
                                row_ex = 'A' + str(non)
                                region = sheet[row_ex].value
                                currentCell = sheet.cell(row=non, column=1)
                                cursor.execute(
                                    "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD'  and Regionname = %s and Period = %s",
                                    (region,period,))
                                result = cursor.fetchone()
                                if result is None:
                                    country = sheet[row_ex].value
                                    currentCell.alignment = Alignment(horizontal='left', indent=2,vertical='bottom',wrap_text=True)
                                    cursor.execute(
                                        "SELECT Value FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Period = %s and Country = %s",
                                        (period,country))
                                    result = cursor.fetchall()
                                    print(result)
                                    if result and len(result) > 0:
                                        val = result[0][0]
                                        sheet.cell(row=non, column=col).value = val

                                    else:
                                        pass
                                else:
                                    for val in result:
                                        sheet.cell(row=non,column=col).value = val
                                        currentCell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
# # #     # # #
#             for col in range(2, max_col):
#                 period = sheet.cell(row=1, column=col).value
#                 print(period)
#                 if (period == '2025-4Q'):
#                     break
#                     # for supply nonopec
#                 total_non_opec_supply = int(fetch.get('total_non_opec_supply')) + 2
#                 print(total_non_opec_supply)
#                 totalnonoecd_supply = int(fetch.get('totalnonoecd_supply'))
#                 print(totalnonoecd_supply)
#                 for nonopec in range(totalnonoecd_supply, total_non_opec_supply):
#                     row_ex = 'A' + str(nonopec)    #
#                     if (period == '2025-4Q'):
#                         flag = 1
#                     progain = sheet[row_ex].value
#                     if sheet[row_ex].value ==  'Processing gains':
#                         print("Processing gainsProcessing gains")
#                         currentCell = sheet.cell(row=nonopec, column=1)
#                         # print(row)
#                         currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                           vertical='bottom',
#                                                           wrap_text=True)
#                         cursor.execute(
#                         "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s and ReportedDate like  '2019-02-13%'",
#                                         (progain,period,))
#                         result = cursor.fetchone()
#                         if result is None:
#                             val = ''
#                             sheet.cell(row=nonopec, column=col).value = val
#
#                         else:
#                             for val in result:
#                                 print(val)
#                                 sheet.cell(row=nonopec, column=col).value = val
#
#
#                     elif sheet[row_ex].value == 'Global Biofuels':
#                         currentCell = sheet.cell(row=nonopec, column=1)
#                         # print(row)
#                         currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                           vertical='bottom',
#                                                           wrap_text=True)
#                         gobal_f=sheet[row_ex].value
#                         cursor.execute(
#                             "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s and ReportedDate like  '2019-02-13%'",
#                             (gobal_f, period,))
#                         result = cursor.fetchone()
#                         #print(result)
#                         if result is None:
#                             val = ''
#                             sheet.cell(row=nonopec, column=col).value = val
#                         else:
#                             for val in result:
#                                 sheet.cell(row=nonopec, column=col).value = val
#                     elif sheet[row_ex].value == 'Total Non-OPEC Supply':
#                         currentCell = sheet.cell(row=nonopec, column=1)
#                         # print(row)
#                         currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                           vertical='bottom',
#                                                           wrap_text=True)
#                         cursor.execute(
#                             "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC' and Period = %s and ReportedDate like  '2019-02-13%'",
#                             ( period,))
#                         result = cursor.fetchone()
#                         if result is None:
#                             val = ''
#                             sheet.cell(row=nonopec, column=col).value = val
#                         else:
#                             for val in result:
#                                 sheet.cell(row=nonopec, column=col).value = val
#                     elif sheet[row_ex].value == 'Non-OPEC: Historical Composition':
#                         currentCell = sheet.cell(row=nonopec, column=1)
#                         # print(row)
#                         currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                           vertical='bottom',
#                                                           wrap_text=True)
#                         cursor.execute(
#                             "SELECT value FROM IEA_HistoricMacroElementsData WHERE ElementType = 'S' and  TableType = '1' and GroupID1 = '4' and GroupID2 = '7' and PeriodID = %s and ReportedDate like  '2019-02-13%'",
#                             ( period,))
#                         result = cursor.fetchone()
#                         if result is None:
#                             val = ''
#                             sheet.cell(row=nonopec, column=col).value = val
#                         else:
#                             for val in result:
#                                 sheet.cell(row=nonopec, column=col).value = val
# # # # #
#     # # # #
#             for col in range(2, max_col):
#                     period = sheet.cell(row=1, column=col).value
#                     print(period)
#                     if (period == '2025-4Q'):
#                         break# #
#         # # # # #for opec supply
#                     total_opec_supply=int(fetch.get('total_opec_supply'))+10
#                     opec_supply = int(fetch.get('opec_supply'))
#                     for opec in range(opec_supply, total_opec_supply):
#                         row_ex = 'A' + str(opec)
#                         if (period == '2025-4Q'):
#                             break
#                         crude = sheet[row_ex].value
#                         if crude == 'Crude':
#                             currentCell = sheet.cell(row=opec, column=1)
#                             # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s ",
#                                 (crude, period,))
#                             result = cursor.fetchone()
#                             #print(result)
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=opec, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=opec, column=col).value = val
#                         elif sheet[row_ex].value == 'NGL':
#                             currentCell = sheet.cell(row=opec, column=1)
#                             # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             ngl = sheet[row_ex].value
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s ",
#                                 (ngl, period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=opec, column=col).value = val
#                             else:
#                                 for val in result:
#                                      sheet.cell(row=opec, column=col).value = val
#                         elif sheet[row_ex].value == 'Total OPEC':
#                             currentCell = sheet.cell(row=opec, column=1)
#                             # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'OPEC' and Regionname IS NULL and Period = %s ",
#                                 ( period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=opec, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=opec, column=col).value = val
#                         elif sheet[row_ex].value == 'OPEC: Historical Composition':
#                             currentCell = sheet.cell(row=opec, column=1)
#                             # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM IEA_HistoricMacroElementsData WHERE ElementType = 'S' and  TableType = '1' and GroupID1 = '3' and GroupID2 = '7' and PeriodID = %s ",
#                                 ( period,))
#                             result = cursor.fetchone()
#                             print("historical OPEC",result)
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=opec, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=opec, column=col).value = val
#                         elif sheet[row_ex].value == 'Total Supply':
#                             currentCell = sheet.cell(row=opec, column=1)
#                             # print(row)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups IS NULL and Period = %s ",
#                                 (period,))
#                             result = cursor.fetchone()
#
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=opec, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=opec, column=col).value = val
#                         else:
#                             row_ex = 'A' + str(opec)
#                             region = sheet[row_ex].value
#                             currentCell = sheet.cell(row=opec, column=1)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'OPEC'  and Regionname = %s and Period = %s",
#                                 (region,period, ))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 country = sheet[row_ex].value
#                                 currentCell.alignment = Alignment(horizontal='left', indent=2, vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT Value FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'OPEC'  and Period = %s and Country = %s ",
#                                     (period, country,))
#                                 result = cursor.fetchall()
#                                 if result and len(result) > 0:
#                                     val = result[0][0]
#                                     sheet.cell(row=opec, column=col).value = val
#
#                                 else:
#                                     pass
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=opec, column=col).value = val
#                                     currentCell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
            wb.save('./IEA QUARTERLY FILE_WITH MODEL_PW_Country.xlsx')
        except Error as e:
            print("print", e)
# # #     # #

    ## To insert Annual data fron tables
    def annual_data():
        # try:
        #     sheet = wb.get_sheet_by_name('Annual_Previous')
        #     wb.remove_sheet(sheet)
        #     sheet = wb.get_sheet_by_name('Annual')
        #     ws2 = wb.copy_worksheet(sheet)
        #     ws2.title = 'Annual_Previous'
        #     wb.remove_sheet(sheet)
        # except Exception as e:
        #     sheet = wb.get_sheet_by_name('Annual')
        #     ws2 = wb.copy_worksheet(sheet)
        #     ws2.title = 'Annual_Previous'
        #     wb.remove_sheet(sheet)

        # region = regions('Annual')
        fetch = fetch_excel("Annual_Previous")
        sheet = wb.get_sheet_by_name('Annual_Previous')
        max_col = (sheet.max_column)
        print(max_col)
        fetch = fetch[0]
        print(fetch)#
    # #     # print(max_row)
        try:
    #             # sheet['A12'].alignment = Alignment(horizontal="left")
    #             for col in range(2, max_col):
    #                     period = sheet.cell(row=1, column=col).value
    #                     print(period)
    #                     if (period == '2025-Y'):
    #                         break
    #                     totaloecd_demand = int(fetch.get('totaloecd_demand')) + 1
    #                     oecd_demand = int(fetch.get('oecd_demand'))
    #                     for row in range(oecd_demand, totaloecd_demand):
    #                         row_ex = 'A' + str(row)
    #                         if (period == '2025-Y'):
    #                             flag = 1
    #                             break
    #                         else:
    #                             if sheet[row_ex].value == 'Total OECD':
    #                                # print("total oecds")
    #                                 currentCell = sheet.cell(row=row, column=1)
    #                                # print(row)
    #                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
    #                                                                  vertical='bottom',
    #                                                                  wrap_text=True)
    #                                 cursor.execute(
    #                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'OECD' and Period = %s and RegionName IS NULL  and ReportedDate <=  '2019-01-18 00:00:00'",
    #                                     (period,))
    #                                 result = cursor.fetchone()
    #                                 if result is None:
    #                                     val = ''
    #                                     sheet.cell(row=row, column=col).value = val
    #                                 else:
    #                                     if ('Y' in period):
    #                                         for val in result:
    #                                             sheet.cell(row=row, column=col).value = val
    #                             else:
    #                                 region = sheet[row_ex].value
    #                                 cursor.execute(
    #                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'OECD' and Period = %s and Regionname = %s and ReportedDate <=  '2019-01-18 00:00:00'",
    #                                     (period, region,))
    #                                 result = cursor.fetchall()
    #                                 if result and len(result) >0:
    #                                     val = result[0][0]
    #                                     print(val)
    #                                     if ('Y' in period):
    #                                         sheet.cell(row=row, column=col).value = val
    #                                         currentCell = sheet.cell(row=row, column=1)
    #                                         # print(row)
    #                                         currentCell.alignment = Alignment(horizontal='left', indent=1, vertical='bottom',
    #                                                                           wrap_text=True)
    #
    #                                 else:
    #                                     country = sheet[row_ex].value
    #                                     cursor.execute(
    #                                         "SELECT value FROM ReportFor2b WHERE ElementType = 'Demand' and Groups = 'OECD' and Period = %s and Country = %s and ReportedDate <=  '2019-01-18 00:00:00'",
    #                                         (period, country,))
    #                                     result = cursor.fetchall()
    #                                     if result and len(result) > 0:
    #                                         val = result[0][0]
    #                                         if ('Y' in period):
    #                                             sheet.cell(row=row, column=col).value = val
    #                                             currentCell = sheet.cell(row=row, column=1)
    #                                             # print(row)
    #                                             currentCell.alignment = Alignment(horizontal='left', indent=2,
    #                                                                               vertical='bottom',
    #                                                                               wrap_text=True)
    #
    #                                     else:
    #                                         pass

# # # #     # #
# # # # #     # #
#                 for col in range(2, max_col):
#                         period = sheet.cell(row=1, column=col).value
#                         print(period)
#                         if (period == '2025-Y'):
#                             break
#         #             #for demand and nonoecd    #
#                         totalnonoecd_demand =  int(fetch.get('totalnonoecd_demand')) + 2
#                         #for total demand
#                         nonoecd_demand = int(fetch.get('nonoecd_demand'))
#                         for non in range(nonoecd_demand, totalnonoecd_demand):
#                             row_ex = 'A' + str(non)
#                             if sheet[row_ex].value == 'Total Non-OECD':
#                                 currentCell = sheet.cell(row=non, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#             #                         #print("total non oecds")
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'Non-OECD' and Period = %s and RegionName IS NULL ",
#                                     (period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=non, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=non, column=col).value = val
#                             else:
#                                 row_ex = 'A' + str(non)
#                                 region = sheet[row_ex].value
#                                 cursor.execute(
#                                     "SELECT Value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups = 'Non-OECD' and Period = %s and RegionName = %s ",
#                                     (period, region,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     country = sheet[row_ex].value
#                                     cursor.execute(
#                                         "SELECT value FROM ReportFor2b WHERE ElementType = 'Demand' and Groups = 'Non-OECD' and Period = %s and Country = %s",
#                                         (period, country,))
#                                     result = cursor.fetchall()
#                                     if result and len(result) > 0:
#                                         val = result[0][0]
#                                         if ('Y' in period):
#                                             sheet.cell(row=non, column=col).value = val
#                                             currentCell = sheet.cell(row=non, column=1)
#                                             # print(row)
#                                             currentCell.alignment = Alignment(horizontal='left', indent=2,
#                                                                               vertical='bottom',
#                                                                               wrap_text=True)
#
#                                     else:
#                                         pass
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=non, column=col).value = val
#                                         currentCell = sheet.cell(row=non, column=1)
#                                         currentCell.alignment = Alignment(horizontal='left', indent=1,
#                                                                           vertical='bottom',
#                                                                           wrap_text=True)
                            # if sheet[row_ex].value == 'Total Demand':
                            #     currentCell = sheet.cell(row=non, column=1)
                            #     # print(row)
                            #     currentCell.alignment = Alignment(horizontal='left', indent=0,
                            #                                       vertical='bottom',
                            #                                       wrap_text=True)
                            #     cursor.execute(
                            #         "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Demand' and Groups IS NULL and Period = %s and RegionName IS NULL ",
                            #         (period,))
                            #     result = cursor.fetchone()
                            #     if result is None:
                            #         val = ''
                            #         sheet.cell(row=non, column=col).value = val
                            #     else:
                            #         for val in result:
                            #             sheet.cell(row=non, column=col).value = val
                            #     break
#     #     # #     #
#                 for col in range(2, max_col):
#                         period = sheet.cell(row=1, column=col).value
#                         print(period)
#                         if (period == '2025-Y'):
#                             break
#                         totaloecd_supply = int(fetch.get('totaloecd_supply')) +1
#                         oecd_supply =  int(fetch.get('oecd_supply'))
#                         for row in range(oecd_supply, totaloecd_supply):
#                             row_ex = 'A' + str(row)
#                             if sheet[row_ex].value == 'Total OECD':
#                                 # print("total supply oecds")
#                                 currentCell = sheet.cell(row=row, column=1)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-OECD' and Period = %s and RegionName IS NULL ",
#                                     (period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=row, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=row, column=col).value = val
#
#                                 break
#                             else:
#                                 row_ex = 'A' + str(row)
#                                 region = sheet[row_ex].value
#                                 currentCell = sheet.cell(row=row, column=1)
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-OECD'  and Regionname = %s and Period = %s ",
#                                     (region,period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     country = sheet[row_ex].value
#                                     currentCell = sheet.cell(row=row, column=1)
#                                     currentCell.alignment = Alignment(horizontal='left', indent=2,
#                                                                       vertical='bottom',
#                                                                       wrap_text=True)
#                                     cursor.execute(
#                                         "SELECT Value FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-OECD'  and Period = %s and Country = %s ",
#                                         (period, country,))
#                                     result = cursor.fetchall()
#                                     if result and len(result) > 0:
#                                         val = result[0][0]
#                                         sheet.cell(row=row, column=col).value = val
#
#                                     else:
#                                         country = sheet[row_ex].value ##for uk
#                                         currentCell.alignment = Alignment(horizontal='left', indent=2,
#                                                                           vertical='bottom', wrap_text=True)
#                                         cursor.execute(
#                                             "SELECT Value FROM Report3 WHERE ElementType = 'Supply' and Period = %s and Country = %s ",
#                                             (period, country,))
#                                         result = cursor.fetchall()
#                                         if result and len(result) > 0:
#                                             val = result[0][0]
#                                             sheet.cell(row=row, column=col).value = val
#
#                                         else:
#                                             pass
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=row, column=col).value = val
#                                         currentCell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
#                                         print(val, period)
# # #     # #     # # #
                for col in range(2, max_col):
                        period = sheet.cell(row=1, column=col).value
                        print(period)
                        if (period == '2025-Y'):
                            break    # for supply non-oecd
                        totalnonoecd_supply = int(fetch.get('totalnonoecd_supply')) + 1
                        nonoecd_supply = int(fetch.get('nonoecd_supply'))
                        for non in range(nonoecd_supply, totalnonoecd_supply):
                            row_ex = 'A' + str(non)
                            if sheet[row_ex].value == 'Total Non-OECD':
                                currentCell = sheet.cell(row=non, column=1)
                                # print(row)
                                currentCell.alignment = Alignment(horizontal='left', indent=0,
                                                                  vertical='bottom',
                                                                  wrap_text=True)
                                cursor.execute(
                                    "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Period = %s and RegionName IS NULL ",
                                    (period,))
                                result = cursor.fetchone()
                                if result is None:
                                    val = ''
                                    sheet.cell(row=non, column=col).value = val
                                else:
                                    for val in result:
                                        sheet.cell(row=non, column=col).value = val
                                break
                            else:
                                row_ex = 'A' + str(non)
                                region = sheet[row_ex].value
                                currentCell = sheet.cell(row=non, column=1)
                                cursor.execute(
                                    "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Regionname = %s and Period = %s ",
                                    (region,period,))
                                result = cursor.fetchone()
                                if result is None:
                                    country = sheet[row_ex].value
                                    currentCell.alignment = Alignment(horizontal='left', indent=2,vertical='bottom',wrap_text=True)
                                    cursor.execute(
                                        "SELECT Value FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Period = %s and Country = %s ",
                                        (period, country,))
                                    result = cursor.fetchall()
                                    if result and len(result) > 0:
                                        val = result[0][0]
                                        sheet.cell(row=non, column=col).value = val

                                    else:
                                        pass
                                else:
                                    for val in result:
                                        sheet.cell(row=non,column=col).value = val
                                        currentCell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
# # # # #     # #     # # # #
#                 for col in range(2, max_col):
#                     period = sheet.cell(row=1, column=col).value
#                     print(period)
#                     if (period == '2025-Y'):
#                         break
#                         # for supply nonopec
#                     total_non_opec_supply = int(fetch.get('total_non_opec_supply'))
#                     totalnonoecd_supply = int(fetch.get('totalnonoecd_supply')) + 1
#                     nonoecd_supply = int(fetch.get('nonoecd_supply'))
#                     for non in range(nonoecd_supply, totalnonoecd_supply):
#                         row_ex = 'A' + str(non)
#                         if sheet[row_ex].value == 'Total Non-OECD':
#                             currentCell = sheet.cell(row=non, column=1)
#                             currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                               vertical='bottom',
#                                                               wrap_text=True)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Period = %s and RegionName IS NULL and ReportedDate like  '2019-02-13%'",
#                                 (period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 val = ''
#                                 sheet.cell(row=non, column=col).value = val
#                             else:
#                                 for val in result:
#                                     sheet.cell(row=non, column=col).value = val
#                             break
#                         else:
#                             row_ex = 'A' + str(non)
#                             region = sheet[row_ex].value
#                             currentCell = sheet.cell(row=non, column=1)
#                             cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Regionname = %s and Period = %s and ReportedDate like  '2019-02-13%'",
#                                 (region,period,))
#                             result = cursor.fetchone()
#                             if result is None:
#                                 country = sheet[row_ex].value
#                                 currentCell.alignment = Alignment(horizontal='left', indent=2,vertical='bottom',wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT Value FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'Non-OPEC-Non-OECD' and Period = %s and Country = %s and ReportedDate like  '2019-02-13%'",
#                                     (period, country,))
#                                 result = cursor.fetchall()
#                                 if result and len(result) > 0:
#                                     val = result[0][0]
#                                     sheet.cell(row=non, column=col).value = val
#
#                                 else:
#                                     pass
#                             else:
#                                 for val in result:
#                                     print(val)
#                                     sheet.cell(row=non,column=col).value = val
#                                     currentCell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
#                     print(total_non_opec_supply)
#                     totalnonoecd_supply = int(fetch.get('totalnonoecd_supply'))
#                     print(totalnonoecd_supply)
#                 for col in range(2, max_col):
#                         period = sheet.cell(row=1, column=col).value
#                         print(period)
#                         if (period == '2025-Y'):
#                             break
#                         for nonopec in range(totalnonoecd_supply, total_non_opec_supply):
#                             row_ex = 'A' + str(nonopec)    #
#                             if (period == '2025-Y'):
#                                 break
#                             progain = sheet[row_ex].value
#                             print(progain)
#                             if sheet[row_ex].value ==  'Processing gains':
#                                 print("Processing gainsProcessing gains")
#                                 currentCell = sheet.cell(row=nonopec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                 "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s and ReportedDate like  '2019-02-13%'",
#                                                 (progain,period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=nonopec, column=col).value = val
#
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=nonopec, column=col).value = val
#
#                             elif sheet[row_ex].value == 'Global Biofuels':
#                                 currentCell = sheet.cell(row=nonopec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 gobal_f=sheet[row_ex].value
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s and ReportedDate like  '2019-02-13%'",
#                                     (gobal_f, period,))
#                                 result = cursor.fetchone()
#                                 #print(result)
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=nonopec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=nonopec, column=col).value = val
#                             elif sheet[row_ex].value == 'Total Non-OPEC Supply':
#                                 currentCell = sheet.cell(row=nonopec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'Non-OPEC' and Period = %s and ReportedDate like  '2019-02-13%'",
#                                     ( period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=nonopec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=nonopec, column=col).value = val
#                             elif sheet[row_ex].value == 'Non-OPEC: Historical Composition':
#                                 currentCell = sheet.cell(row=nonopec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT value FROM IEA_HistoricMacroElementsData WHERE ElementType = 'S' and  TableType = '1' and GroupID1 = '4' and GroupID2 = '7' and PeriodID = %s and ReportedDate like  '2019-02-13%'",
#                                     ( period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=nonopec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=nonopec, column=col).value = val
    # #     # #
# #         # # # #
#                 for col in range(2, max_col):
#                         period = sheet.cell(row=1, column=col).value
#                         print(period)
#                         if (period == '2025-Y'):
#                             break# #
#             # # # # #for opec supply
#                         total_opec_supply=int(fetch.get('total_opec_supply'))+10
#                         opec_supply = int(fetch.get('opec_supply'))
#                         for opec in range(opec_supply, total_opec_supply):
#                             row_ex = 'A' + str(opec)
#                             if (period == '2025-Y'):
#                                 break
#                             crude = sheet[row_ex].value
#                             if crude == 'Crude':
#                                 currentCell = sheet.cell(row=opec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s ",
#                                     (crude, period,))
#                                 result = cursor.fetchone()
#                                 #print(result)
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=opec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=opec, column=col).value = val
#                             elif sheet[row_ex].value == 'NGL':
#                                 currentCell = sheet.cell(row=opec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 ngl = sheet[row_ex].value
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = %s  and Period = %s ",
#                                     (ngl, period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=opec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                          sheet.cell(row=opec, column=col).value = val
#                             elif sheet[row_ex].value == 'Total OPEC':
#                                 currentCell = sheet.cell(row=opec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'OPEC' and RegionName IS NULL and Period = %s",
#                                     ( period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=opec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=opec, column=col).value = val
#                             elif sheet[row_ex].value == 'OPEC: Historical Composition':
#                                 currentCell = sheet.cell(row=opec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT value FROM IEA_HistoricMacroElementsData WHERE ElementType = 'S' and  TableType = '1' and GroupID1 = '3' and GroupID2 = '7' and PeriodID = %s",
#                                     ( period,))
#                                 result = cursor.fetchone()
#                                 print("historical OPEC",result)
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=opec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=opec, column=col).value = val
#                             elif sheet[row_ex].value == 'Total Supply':
#                                 currentCell = sheet.cell(row=opec, column=1)
#                                 # print(row)
#                                 currentCell.alignment = Alignment(horizontal='left', indent=0,
#                                                                   vertical='bottom',
#                                                                   wrap_text=True)
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups IS NULL and Period = %s",
#                                     (period,))
#                                 result = cursor.fetchone()
#
#                                 if result is None:
#                                     val = ''
#                                     sheet.cell(row=opec, column=col).value = val
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=opec, column=col).value = val
#                             else:
#                                 row_ex = 'A' + str(opec)
#                                 region = sheet[row_ex].value
#                                 currentCell = sheet.cell(row=opec, column=1)
#                                 cursor.execute(
#                                     "SELECT value FROM TotalSummaryReport WHERE ElementType = 'Supply' and Groups = 'OPEC'  and Regionname = %s and Period = %s",
#                                     (region,period,))
#                                 result = cursor.fetchone()
#                                 if result is None:
#                                     country = sheet[row_ex].value
#                                     currentCell.alignment = Alignment(horizontal='left', indent=2, vertical='bottom',
#                                                                       wrap_text=True)
#                                     cursor.execute(
#                                         "SELECT Value FROM ReportFor3 WHERE ElementType = 'Supply' and Groups = 'OPEC'  and Period = %s and Country = %s",
#                                         (period, country,))
#                                     result = cursor.fetchall()
#                                     if result and len(result) > 0:
#                                         val = result[0][0]
#                                         sheet.cell(row=opec, column=col).value = val
#
#                                     else:
#                                         pass
#                                 else:
#                                     for val in result:
#                                         sheet.cell(row=opec, column=col).value = val
#                                         currentCell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
                wb.save('./IEA QUARTERLY FILE_WITH MODEL_PW_Country.xlsx')
        except Error as e:
            print("print", e)
#
#
    # def delta(wb):  #### to find difference between values of two sheets
    #     print(wb.get_sheet_names())
    #     # sheet = wb.get_sheet_by_name('Delta_annual')
    #     # wb.remove_sheet(sheet)
    #     # sheet = wb.get_sheet_by_name('Delta_Quarter')
    #     # wb.remove_sheet(sheet)
    #     try:
    #         if 'Delta_annual' in wb.sheetnames:
    #             print('sheet1 exists')
    #         else:
    #             sheet = wb.get_sheet_by_name('Annual')
    #             ws2 = wb.copy_worksheet(sheet)
    #             ws2.title = 'Delta_annual'
    #             max_col = (ws2.max_column)
    #             print(max_col)
    #             max_row = (ws2.max_row)
    #             print(max_row)
    #             sheet_copy = wb.get_sheet_by_name('Annual_Previous')
    #             # sheet_copy.cell(row=11, column=2).value = '45'
    #             for col in range(2, max_col):
    #                 period = ws2.cell(row=1, column=col).value
    #                 if (period == '2025-Y'):
    #                     break
    #                 for row in range(11, 112):
    #                     value1 = sheet.cell(row=row, column=col).value
    #                     value2 = sheet_copy.cell(row=row, column=col).value
    #                     print(value1, value2)
    #                     if value1 == None or value2 == None:
    #                         pass
    #                     else:
    #                         try:
    #                             val = int(value1) - int(value2)
    #                             ws2.cell(row=row, column=col).value = val
    #                         except ValueError:
    #                             pass
    #     except Exception as e:
    #         print(str(e))
        # try:
        #     if 'Delta_Quarter' in wb.sheetnames:
        #         print('sheet2 exists')
        #     else:
        #         sheet = wb.get_sheet_by_name('Quarter')
        #         ws2 = wb.copy_worksheet(sheet)
        #         ws2.title = 'Delta_Quarter'
        #         max_col = (ws2.max_column)
        #         print(max_col)
        #         max_row = (ws2.max_row)
        #         print(max_row)
        #         sheet_copy = wb.get_sheet_by_name('Quarter_Previous')
        #         for col in range(2, max_col):
        #             period = ws2.cell(row=1, column=col).value
        #             if (period == '2025-Y'):
        #                 break
        #             for row in range(11, 112):
        #                 value1 = sheet.cell(row=row, column=col).value
        #                 value2 = sheet_copy.cell(row=row, column=col).value
        #                 print(value1, value2)
        #                 if value1 == None or value2 == None:
        #                     pass
        #                 else:
        #                     try:
        #                         val = int(value1) - int(value2)
        #                         ws2.cell(row=row, column=col).value = val
        #                     except ValueError:
        #                         pass
        # except Exception as e:
        #     print(str(e))
        wb.save('./IEA QUARTERLY FILE_WITH MODEL_PW_Country.xlsx')


    def match_excel():
        fetch = fetch_excel('Annual')
        match_fetch = fetch[1]
        print(match_fetch)
        sheet = wb.get_sheet_by_name('Annual')
        max_col = (sheet.max_column)
        print(max_col)
        max_row = (sheet.max_row)
        db_result = {}
        cursor.execute(
            "select * from ExcelMapping")
        result = cursor.fetchall()
        # print(result)
        for results in result:
            # print(results[0])
            db_result[results[0]] =results[1]
        print(db_result)
        #
        for key in db_result:
            if key in match_fetch:
                if (db_result[key] == match_fetch[key]):
                    print(key)
                    value = db_result[key]

                else:
                    print(key)
                    f = open("error_log.txt","a")
                    f.write("error in row")
                    f.write(key)
                    value = db_result[key]
                    f.write(str(value))
                    f.write('\n')


    #     #
    # callproc_script()
    # regions("Annual")
    # match_excel()
    # annual_data()
    quarter_data()
    # quarter_data()

    # wb.save('./IEA QUARTERLY FILE_WITH MODEL_PW_country.xlsx')
    # print("anual data goessss")
    # annual_data()

    # delta(wb)
    # fetch = fetch_excel('Quarter')
    # print(fetch)
except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")










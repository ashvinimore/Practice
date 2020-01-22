import mysql.connector
from openpyxl import load_workbook
from datetime import date


#
# def colnum_string(n):
#     string = ""
#     while n > 0:
#         n, remainder = divmod(n - 1, 26)
#         string = chr(65 + remainder) + string
#     return string

def dumpdataStarSchema(tablename):
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))
            #
            sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            cursor.execute(sqlquery)
            # file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/CDEV_Q2 19_ V111.xlsx',data_only=True)
            sheet = wb.get_sheet_by_name('Industry Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print("max:",max_row,max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0

            try:
                for row in range(1,max_row):
                    for col in range(1,max_col):
                        if sheet.cell(row = row,column = col).value == 'GXBRL ID' or sheet.cell(row = row,column = col).value == 'GXbrleleID':
                          gxbrlid = col
                        if sheet.cell(row = row,column = col).value == 'SegmentID':
                            SegmentID = col
                        if sheet.cell(row = row,column = col).value == 'GICSID':
                            GICSID = col
                        if sheet.cell(row = row,column = col).value == 'Geography':
                            Geography = col
                        if sheet.cell(row = row,column = col).value == 'ProductID':
                            ProductID = col
                        if(sheet.cell(row = row,column = col).value == 'Description'):
                            periodrow = row
                            periodcol = col

                # print(gxbrlid, SegmentID, GICSID, Geography, ProductID, Description)

                #extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])
                print(periodcolumn)

                #for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where Date = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)



                for rowall in range(1, max_row):
                    if (gxbrlid == 0):
                        gxbrlvalue = 'None'
                        segmentvalue = 'None'
                        geographyvalue = 'None'
                        GICSvalue = 'None'
                        Productvalue = 'None'
                    else:
                        gxbrlvalue = str(sheet.cell(row = rowall,column = gxbrlid ).value)
                        print(gxbrlvalue)
                        if (SegmentID == 0):
                            segmentvalue = 'None'
                        else:
                            segmentvalue = sheet.cell(row=rowall, column=SegmentID).value
                            if segmentvalue == None:
                                segmentvalue = 'None'
                        if (Geography == 0):
                            geographyvalue = 'None'
                        else:
                            geographyvalue = sheet.cell(row=rowall, column=Geography).value
                            if (geographyvalue == None):
                                geographyvalue = 'None'
                        if (GICSID == 0):
                            GICSvalue = 'None'
                        else:
                            GICSvalue = sheet.cell(row=rowall, column=GICSID).value
                            if (GICSvalue == None):
                                GICSvalue = 'None'
                        if (ProductID == 0):
                            Productvalue = 'None'
                        else:
                            Productvalue = sheet.cell(row=rowall, column=ProductID).value
                            if (Productvalue == None):
                                Productvalue = 'None'

                        try:
                            if(gxbrlvalue.isdigit() == True):
                                # print(rowall)
                                try:
                                    cursor.execute("SELECT UID FROM  DWHMainQuery WHERE  GXbrlEleID = %s and SegmentID = %s and ProductID = %s and GeoID = %s and GicId = %s",(gxbrlvalue,segmentvalue,Productvalue,geographyvalue,GICSvalue))
                                    uid = cursor.fetchone()[0]
                                    if int(uid) > 0:
                                        # print(uid, gxbrlvalue)
                                        periodconvert = 0
                                        periodconvert = periodcol + 1
                                        # print("periodcolumn",periodconvert,rowall)
                                        for colperiod in range(periodconvert, max_col):
                                            # print(rowall)
                                            value = (sheet.cell(row = rowall,column = colperiod).value)
                                            # print(value)
                                            # print(periodcolumn, rowall,sheet.cell(row = rowall,column = periodcolumn).value)
                                            if int(value) > 0:
                                                periodvalue = (sheet.cell(row = periodrow,column = colperiod).value)
                                                periodvalue = periodvalue.replace(' ','')
                                                # print(periodvalue)

                                                try:
                                                    vperiods = str(repr(periodcolumn).replace('[', '(').replace(']', ')').replace("'",''))
                                                    # print(vperiods)
                                                    periodsearchquery = "SELECT PeriodIDPri FROM CalendarPeriod WHERE %s in {}".format(vperiods)
                                                    cursor.execute(periodsearchquery, (periodvalue,))
                                                    periodid = cursor.fetchall()
                                                    periodid = periodid[0][0]
                                                    # print(value,periodvalue,periodid)
                                                    factidquery = "select count(*) from StarSchema1";
                                                    cursor.execute((factidquery))
                                                    factid = cursor.fetchall()[0][0]
                                                    print(factid)
                                                    if periodid:
                                                        try:
                                                            searchQuery = "select  factid from StarSchema1 where UID = %s and PeriodID = %s and ReportedDateID = %s"
                                                            cursor.execute(searchQuery, (uid,periodid,reporteddate))
                                                            uidexist = cursor.fetchall()
                                                            if uidexist:
                                                                print("uidd")
                                                            else:
                                                                factid = factid + 1
                                                                insertQuery = "insert into StarSchema1(FactID,UID,PeriodID,ReportedDateID,Value) values(" \
                                                                              "'%s','%s','%s','%s','%s')" % (
                                                                                  (factid,uid, periodid, reporteddate,value) )
                                                                try:
                                                                    cursor.execute(insertQuery)
                                                                    print("query excuted successfully")
                                                                    connection.commit()
                                                                except Exception as e:
                                                                    print(e)
                                                        except Exception as e:
                                                            print(e)


                                                except Exception as e:
                                                    print(e,periodvalue)
                                                    pass

                                except Exception as e:
                                    pass
                        except Exception as e:
                            print(e)
                            pass



            except Exception as e:
                print(e)

    except Exception as e:
        print(e)



def dumpPeriodSchema(tablename):
    print("dump uid based schema",tablename)
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            cursor.execute(sqlquery)
            file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/Data Warehouse/DWH-1.xlsx',data_only=True)
            sheet = wb.get_sheet_by_name('Global Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print(max_row,max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0
            factid = 544
            try:
                for row in range(1, max_row):
                    for col in range(1, max_col):
                        if (sheet.cell(row=row, column=col).value) == 'GXBRL ID':
                            gxbrlid = col
                        if sheet.cell(row=row, column=col).value == 'SegmentID':
                            SegmentID = col
                        if sheet.cell(row=row, column=col).value == 'GICSID':
                            GICSID = col
                        if sheet.cell(row=row, column=col).value == 'Geography':
                            Geography = col
                        if sheet.cell(row=row, column=col).value == 'ProductID':
                            ProductID = col
                        if (sheet.cell(row=row, column=col).value == 'Description'):
                            periodrow = row
                            periodcol = col

                # print(gxbrlid, SegmentID, GICSID, Geography, ProductID, Description)

                # extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])

                # for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where Date = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)

                for rowall in range(1, max_row):
                    if (gxbrlid == 0):
                        gxbrlvalue = 'None'
                        segmentvalue = 'None'
                        geographyvalue = 'None'
                        GICSvalue = 'None'
                        Productvalue = 'None'
                    else:
                        gxbrlvalue = str(sheet.cell(row=rowall, column=gxbrlid).value)
                        # print(gxbrlvalue)
                        if (SegmentID == 0):
                            segmentvalue = 'None'
                        else:
                            segmentvalue = sheet.cell(row=rowall, column=SegmentID).value
                            if segmentvalue == None:
                                segmentvalue = 'None'
                        if (Geography == 0):
                            geographyvalue = 'None'
                        else:
                            geographyvalue = sheet.cell(row=rowall, column=Geography).value
                            if (geographyvalue == None):
                                geographyvalue = 'None'
                        if (GICSID == 0):
                            GICSvalue = 'None'
                        else:
                            GICSvalue = sheet.cell(row=rowall, column=GICSID).value
                            if (GICSvalue == None):
                                GICSvalue = 'None'
                        if (ProductID == 0):
                            Productvalue = 'None'
                        else:
                            Productvalue = sheet.cell(row=rowall, column=ProductID).value
                            if (Productvalue == None):
                                Productvalue = 'None'

                        try:
                            if (gxbrlvalue.isdigit() == True):
                                # print(rowall)
                                try:
                                    cursor.execute(
                                        "SELECT UID FROM  DWHMainQuery WHERE  GXbrlEleID = %s and SegmentID = %s and ProductID = %s and GeoID = %s and GicId = %s",
                                        (gxbrlvalue, segmentvalue, Productvalue, geographyvalue, GICSvalue))
                                    uid = cursor.fetchone()[0]
                                    if int(uid) > 0:
                                        # print(uid, gxbrlvalue)
                                        periodconvert = 0
                                        periodconvert = periodcol + 1
                                        # print("periodcolumn",periodconvert,rowall)
                                        for colperiod in range(periodconvert, max_col):
                                            # print(rowall)
                                            value = (sheet.cell(row=rowall, column=colperiod).value)
                                            # print(value)
                                            # print(periodcolumn, rowall,sheet.cell(row = rowall,column = periodcolumn).value)
                                            if int(value) > 0:
                                                periodvalue = (sheet.cell(row=periodrow, column=colperiod).value)
                                                periodvalue = periodvalue.replace(' ', '')
                                                # print(periodvalue)

                                                try:
                                                    vperiods = str(
                                                        repr(periodcolumn).replace('[', '(').replace(']', ')').replace(
                                                            "'", ''))
                                                    # print(vperiods)
                                                    periodsearchquery = "SELECT PeriodIDPri FROM CalendarPeriod WHERE %s in {}".format(
                                                        vperiods)
                                                    cursor.execute(periodsearchquery, (periodvalue,))
                                                    periodid = cursor.fetchall()
                                                    periodid = periodid[0][0]
                                                    # print(value,periodvalue,periodid)
                                                    tablename1 = 'PeriodwiseSchema1'
                                                    if periodid:
                                                        # print(periodid)
                                                        try:
                                                            columnnameexistquery = "select * from information_schema.columns where column_name = %s and TABLE_NAME = %s"
                                                            cursor.execute(columnnameexistquery, (periodid,tablename1,))
                                                            exitstedperiodid =  cursor.fetchall()
                                                            exitstedperiodid = (exitstedperiodid[0][0])
                                                            factidquery = "select  count(*) from PeriodwiseSchema1";
                                                            cursor.execute(factidquery)
                                                            factid= cursor.fetchall()[0][0]
                                                            print(factid)
                                                            if exitstedperiodid:
                                                                try:
                                                                    uidsearch ="select  factid from PeriodwiseSchema1 where DWHID = %s";
                                                                    cursor.execute(uidsearch,
                                                                                   (uid, ))
                                                                    uidexist = cursor.fetchall()[0][0]
                                                                    updatequery = "update PeriodwiseSchema1 set `%s` = %s where DWHID = %s"
                                                                    cursor.execute(updatequery,(periodid,value,uid,))
                                                                    print("query excuted successfully")
                                                                    connection.commit()
                                                                except Exception as e:
                                                                    print(e)

                                                                except IndexError:
                                                                    factid = factid + 1
                                                                    insertQuery = "insert into PeriodwiseSchema1(FactID,ReportedDateID,DWHID,`%s`) values(" \
                                                                                              "'%s','%s','%s','%s')" % (
                                                                                              (periodid,factid, reporteddate,uid,value))
                                                                    try:
                                                                        cursor.execute(insertQuery)
                                                                        print("query excuted successfully")
                                                                        connection.commit()
                                                                    except Exception as e:
                                                                        print(e)

                                                        except IndexError:
                                                            # print(periodid)
                                                            addperiodcolumnquery = "alter table `PeriodwiseSchema1`   add  `%s`  smallint(6) "
                                                            cursor.execute(addperiodcolumnquery, (periodid,))
                                                            print("query excuted successfully")
                                                            connection.commit()
                                                        # try:
                                                        #     searchQuery = "select  factid from PeriodwiseSchema1 where UID = %s and ReportedDateID = %s and %s IS NOT NULL "
                                                        #     cursor.execute(searchQuery, (uid, reporteddate,periodid,))
                                                        #     uidexist = cursor.fetchall()
                                                        #     if uidexist:
                                                        #         print("existsss")
                                                        #     else:
                                                        #         print(value)                                                    #
                                                        #         factid = factid + 1
                                                        #         insertQuery = "insert into PeriodwiseSchema1(FactID,ReportedDateID,UID,`%s`) values(" \
                                                        #                       "'%s','%s','%s','%s')" % (
                                                        #                           (periodid,factid, reporteddate,uid,value))
                                                        #         try:
                                                        #             cursor.execute(insertQuery)
                                                        #             print("query excuted successfully")
                                                        #             connection.commit()
                                                        #         except Exception as e:
                                                        #             print(e)
                                                        except Exception as e:
                                                            print(e)


                                                except Exception as e:
                                                    print(e, periodvalue)
                                                    pass



                                except Exception as e:
                                    pass
                        except Exception as e:
                            print(e)
                            pass



            except Exception as e:
                print(e)




            # addperiodcolumnquery = "alter table  `PeriodwiseSchema1` add column `PeriodId`   smallint(6) "
    except Exception as e:
                print(e)



def dumpUIDSchema(tablename):
    print("dump uid based schema",tablename)
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            sqlquery = "CREATE TABLE IF NOT EXISTS %s" % tablename
            cursor.execute(sqlquery)
            file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/Data Warehouse/DWH-1.xlsx', data_only=True)
            sheet = wb.get_sheet_by_name('Global Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print(max_row, max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0
            # factid = 496
            try:
                for row in range(1, max_row):
                    for col in range(1, max_col):
                        if (sheet.cell(row=row, column=col).value) == 'GXBRL ID':
                            gxbrlid = col
                        if sheet.cell(row=row, column=col).value == 'SegmentID':
                            SegmentID = col
                        if sheet.cell(row=row, column=col).value == 'GICSID':
                            GICSID = col
                        if sheet.cell(row=row, column=col).value == 'Geography':
                            Geography = col
                        if sheet.cell(row=row, column=col).value == 'ProductID':
                            ProductID = col
                        if (sheet.cell(row=row, column=col).value == 'Description'):
                            periodrow = row
                            periodcol = col

                # print(gxbrlid, SegmentID, GICSID, Geography, ProductID, Description)

                # extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])

                # for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where Date = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)

                for rowall in range(1, max_row):
                    if (gxbrlid == 0):
                        gxbrlvalue = 'None'
                        segmentvalue = 'None'
                        geographyvalue = 'None'
                        GICSvalue = 'None'
                        Productvalue = 'None'
                    else:
                        gxbrlvalue = str(sheet.cell(row=rowall, column=gxbrlid).value)
                        # print(gxbrlvalue)
                        if (SegmentID == 0):
                            segmentvalue = 'None'
                        else:
                            segmentvalue = sheet.cell(row=rowall, column=SegmentID).value
                            if segmentvalue == None:
                                segmentvalue = 'None'
                        if (Geography == 0):
                            geographyvalue = 'None'
                        else:
                            geographyvalue = sheet.cell(row=rowall, column=Geography).value
                            if (geographyvalue == None):
                                geographyvalue = 'None'
                        if (GICSID == 0):
                            GICSvalue = 'None'
                        else:
                            GICSvalue = sheet.cell(row=rowall, column=GICSID).value
                            if (GICSvalue == None):
                                GICSvalue = 'None'
                        if (ProductID == 0):
                            Productvalue = 'None'
                        else:
                            Productvalue = sheet.cell(row=rowall, column=ProductID).value
                            if (Productvalue == None):
                                Productvalue = 'None'
                #
                        try:
                            tablename1 = 'UIDwiseSchema1'
                            if (gxbrlvalue.isdigit() == True):
                                try:
                                    cursor.execute(
                                        "SELECT UID FROM  DWHMainQuery WHERE  GXbrlEleID = %s and SegmentID = %s and ProductID = %s and GeoID = %s and GicId = %s",
                                        (gxbrlvalue, segmentvalue, Productvalue, geographyvalue, GICSvalue))
                                    uid = cursor.fetchone()[0]
                                    # print(uid)
                                    if int(uid) > 0:
                                        try:
                                            columnnameexistquery = "select * from information_schema.columns where column_name = %s and TABLE_NAME = %s"
                                            cursor.execute(columnnameexistquery, (uid, tablename1,))
                                            exitsteduidid = cursor.fetchall()
                                            exitsteduidid = (exitsteduidid[0][0])
                                            # print("exitsteduidid",exitsteduidid)#                                     #

                                        except IndexError:
                                            # print("eror")
                                            adduidcolumnquery = "alter table `UIDwiseSchema1`add  `%s`  smallint(6) "
                                            cursor.execute(adduidcolumnquery, (uid,))
                                            print("query excuted successfully")
                                            connection.commit()

                                    periodconvert = periodcol + 1

                                    for colperiod in range(periodconvert, max_col):
                                        periodvalue = (sheet.cell(row=periodrow, column=colperiod).value)
                                        periodvalue = periodvalue.replace(' ', '')
                                        value  = (sheet.cell(row=rowall, column=colperiod).value)
                                        if value > 0:
                                            try:
                                                vperiods = str(
                                                    repr(periodcolumn).replace('[', '(').replace(']', ')').replace(
                                                        "'", ''))
                                                periodsearchquery = "SELECT PeriodIDPri FROM CalendarPeriod WHERE %s in {}".format(
                                                    vperiods)
                                                cursor.execute(periodsearchquery, (periodvalue,))
                                                periodid = cursor.fetchall()
                                                periodid = periodid[0][0]
                                                countfactid = "select count(*) from UIDwiseSchema1"
                                                cursor.execute(countfactid, )
                                                try:
                                                    countfactid = cursor.fetchall()[0][0]
                                                except Exception as e:
                                                    countfactid = 0
                                                periodsearch = "select factid from UIDwiseSchema1 where PeriodID = %s"
                                                cursor.execute(periodsearch, (periodid,))
                                                factid = cursor.fetchall()
                                                if (len(factid) >= 1):
                                                    updatevalue = "Update UIDwiseSchema1 set `%s` = %s where PeriodID = %s"
                                                    try:
                                                        cursor.execute(updatevalue,(uid,value,periodid,))
                                                        print("query excuted successfully")
                                                        connection.commit()
                                                    except Exception as e:
                                                        print("insert", e)

                                                else:
                                                    countfactid = countfactid + 1
                                                    insertQuery = "insert into UIDwiseSchema1(FactID,PeriodID,ReportedDateID    )  values(" \
                                                                  "'%s','%s','%s')" % (
                                                                      (countfactid,
                                                                       periodid, reporteddate,
                                                                       )
                                                                  )
                                                    try:
                                                        cursor.execute(insertQuery)
                                                        print("query excuted successfully")
                                                        connection.commit()
                                                    except Exception as e:
                                                        print("insert", e)

                                            except Exception as e:
                                                print(e)

                                except Exception as e:
                                    pass
                        except Exception as e:
                            print(e)
                            pass
            except Exception as e:
                print(e)

    except Exception as e:
        print(e)



tablename = "`StarSchema1` (`FactID` smallint(6) NOT NULL, `UID` smallint(6) NOT NULL,  `PeriodID` varchar(50) NOT NULL,  `ReportedDateID` varchar(256) DEFAULT NULL,  `Value` int(11) DEFAULT NULL,  PRIMARY KEY (`FactID`)) ENGINE=InnoDB DEFAULT CHARSET=latin1;"
dumpdataStarSchema(tablename)
tablename = "`PeriodwiseSchema1` (`FactID` smallint(6) NOT NULL,`ReportedDateID` varchar(256) DEFAULT NULL,`DWHID` int (23), PRIMARY KEY (`FactID`)) ENGINE=InnoDB DEFAULT CHARSET=latin1;"
dumpPeriodSchema(tablename)
# tablename = "`UIDwiseSchema1` (`FactID` smallint(6) NOT NULL,`PeriodID` varchar(50) NOT NULL, `ReportedDateID` varchar(256) DEFAULT NULL,  PRIMARY KEY (`FactID`)) ENGINE=InnoDB DEFAULT CHARSET=latin1;"
# dumpUIDSchema (tablename)




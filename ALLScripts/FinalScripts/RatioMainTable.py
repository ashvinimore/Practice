import mysql.connector
from openpyxl import load_workbook
from datetime import date
try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='DWHDB',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
        # result = cursor.execute("select * from DWHMainQuery")
        # print(result)

        wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/SLB_Q2 19.xlsx',data_only=True)
        ticker = 'SLB'
        coid = "SELECT CoID FROM DeCompany where Ticker = '%s'"%ticker
        cursor.execute(coid,)
        result = cursor.fetchall()
        result = str(result[0][0])
        sheet = wb.get_sheet_by_name('Ratios')
        max_col = (sheet.max_column)
        # print(max_col)
        max_row = (sheet.max_row)
        priid = 0
        for row in range(1,max_row):
            ratioid = 'A' + str(row)
            ratiovalue = str(sheet[ratioid].value)
            if sheet[ratioid].value:
                if (ratiovalue).isnumeric() == True:
                   if sheet[ratioid].value <= 9:
                       priid = result+'0'+ratiovalue

                   else:
                       priid = result + ratiovalue
                   print(ratiovalue, result)
                   findQuery = "select * from DWHRatioMainQuery where DWHRatioID = %s and DWHCoID = %s"
                   cursor.execute(findQuery, (ratiovalue, result,))
                   results = cursor.fetchall()
                   print(priid, ratiovalue, result, results)
                   if results == []:
                       insertQuery = "INSERT INTO DWHRatioMainQuery(DWHUID,DWHRatioID,DWHCoID) VALUES (%s,%s,%s)"
                       cursor.execute(insertQuery,(priid, ratiovalue,result,))
                       connection.commit()
                   else:
                       pass



except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")
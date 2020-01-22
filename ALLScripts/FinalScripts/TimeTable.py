# import pandas as pd
# datelist = pd.date_range(pd.datetime.today(), periods=10000).tolist()
# for val in datelist:
#    val = str(val)
#    print(val.split(' ')[0])

import mysql.connector
from openpyxl import load_workbook
import time
from calendar import timegm


import datetime
try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='DWHDB',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
        file_name = '/home/pavan/Music/Data Warehouse/Period/period.xlsx'
        wb = load_workbook('/home/pavan/Music/Data Warehouse/Period/period.xlsx')
        # print(wb.get_sheet_names())

        sheet = wb.get_sheet_by_name('period')
        max_col = (sheet.max_column)
        # print(max_col)
        max_row = (sheet.max_row) + 1
        # print(max_row)
        for row in range(2, max_row):
            row_extime = 'B' + str(row)
            time1 = ( sheet[row_extime].value)
            times = (time1.split(":"))
            utc_time = time.strptime("1970-01-01T"+times[0] + ":" + times[1]+":" + times[2] , "%Y-%m-%dT%H:%M:%S")
            epoch_time = timegm(utc_time)
            print(epoch_time)
            insertQuery = "insert into CalendarTime(TimeID,Time)  values(""'%s','%s')" % (
            (epoch_time, time1,))
            try:
                cursor.execute(insertQuery)
                print("query excuted successfully")
                connection.commit()
            except Exception as e:
                print(e)


    # start = datetime.datetime.strptime("01-01-1995", "%d-%m-%Y")
    # end = datetime.datetime.strptime("31-12-2050", "%d-%m-%Y")
    # date_generated = [start + datetime.timedelta(days=x) for x in range(0, (end-start).days)]
    #
    # for date in date_generated:
    #     yydate =  (date.strftime("%Y-%m-%d"))
    #     standarddate = str (date.strftime("%d-%m-%Y"))
    #     dateid =  (date.strftime("%Y%m%d"))
    #     print(dateid)
    #     insertQuery = "insert into CalendarDate(DateID,Date,StandardDateFormat)  values(""'%s','%s','%s')" % ((dateid, yydate, standarddate,))
    #     try:
    #         cursor.execute(insertQuery)
    #         print("query excuted successfully")
    #         connection.commit()
    #     except Exception as e:
    #         print(e)



except Exception as e:
    print("Error while connecting to MySQL",e)


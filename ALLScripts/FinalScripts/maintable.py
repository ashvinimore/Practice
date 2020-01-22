import os
import re
import csv
import mysql.connector
from openpyxl import load_workbook

try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='DWHDB',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
        # result = cursor.execute("select * from DWHMainQuery")
        # print(result)
        file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
        wb = load_workbook('/home/pavan/Music/DataWarehouse/DWH-1.xlsx')
        # print(wb.get_sheet_names())
        # anotherSheet = wb.active
        #
        # match_fetch = {}
        sheet = wb.get_sheet_by_name('Global Template')
        max_col = (sheet.max_column)
        # print(max_col)
        max_row = (sheet.max_row)
        autoid = 1
        # print(max_row)
        try:
            result = []
            for row in range(1, max_row):
                row_ex = 'L' + str(row)
                if ( sheet[row_ex].value == 'GXBRL ID' or  sheet[row_ex].value == 'GXbrleleID'):
                    # print(row)
                    start = row + 1
                    for val in range(start,max_row):
                        rowval = 'L' + str(val)
                        rowseg = 'Q' + str(val)
                        roegic = 'T'+ str(val)
                        rowprod = 'E' + str(val)
                        rowgeo = 'S'+ str(val)
                        gxbrl = sheet[rowval].value
                        segments = sheet[rowseg].value
                        gicid = sheet[roegic].value
                        prodid = sheet[rowprod].value
                        geoid = sheet[rowgeo].value

                        if(gxbrl != None or segments != None  or gicid != None or geoid != None) :
                            if (gxbrl == None):
                                gxbrl = None
                            if (segments == None):
                                segments = None

                            if (gicid == None):
                                gicid = None

                            if (geoid == None):
                                geoid = None

                            if (prodid == None):
                                prodid = None


                            cursor.execute(
                                "SELECT * FROM  DWHMainQuery WHERE  GXbrlEleID = %s and SegmentID = %s and GeoID = %s and GicID = %s and ProductID = %s",
                                (gxbrl,segments,geoid,gicid,prodid))
                            result = cursor.fetchall()

                            if(result == []):
                                parent = None
                                autoid = autoid + 1
                                UID = autoid
                                print(UID,gxbrl,segments,geoid,gicid,prodid)
                                insertQuery = "insert into `DWHMainQuery`(DWHID,GXbrlEleID,SegmentID,GeoID,GicID,ProductID)  values(%s,%s,%s,%s,%s,%s)"
                                # insertQuery = "insert into GlobalElement(GXbrlEleID,Description) values('1','1')"
                                try:
                                    cursor.execute(insertQuery,((UID,gxbrl,segments,geoid,gicid,prodid,)))
                                    print("query excuted successfully")
                                    connection.commit()
                                except Exception as e:
                                    print(e)
                                    pass

                            #
                            #
                            else:
                                print("already ")

                        # if()
                # for value in range(0,13):
                #    reader.pop(value)
                # for row in reader:

        except Exception as e:
            raise e

except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")
# import pandas as pd
# datelist = pd.date_range(pd.datetime.today(), periods=10000).tolist()
# for val in datelist:
#    val = str(val)
#    print(val.split(' ')[0])

import mysql.connector
from openpyxl import load_workbook


import datetime
def periods(periodss):
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            record = cursor.fetchone()
            print("Your connected to - ", record)
            # file_name = '/home/pavan/Music/Data Warehouse/Period/period.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/PeriodFinal1.xlsx')
            # print(wb.get_sheet_names())

            # sheet = wb.get_sheet_by_name('Period')
            # max_col = (sheet.max_column)
            # max_row = (sheet.max_row) + 1
            for periodyear in periodss:
                selectQuery = "select PeriodId from CalendarPeriodNew"
                cursor.execute(selectQuery)
                result = cursor.fetchall()

                for re in result :
                    print(re[0])
                    periodidpri = 'P' + str(re[0])
                    try:
                        UpdateQuery  = "update CalendarPeriodNew set PeriodSchemaID = '%s' where PeriodID = %s"%(periodidpri,str(re[0]))
                        cursor.execute(UpdateQuery)
                        connection.commit()
                        print("updated")
                    except Exception as e:
                        print(e)
                #     if(re[2] <= 9):
                #        modeid = '0' + str(re[2])
                #     else:
                #         modeid = str(re[2])
                #     if(re[3] <= 9):
                #         idx= '0' + str(re[3])
                #     else:
                #         idx = str(re[3])
                #     periodidpri = str(periodyear) + modeid + idx
                #     print(re[0],periodidpri)
                #     try:
                #         UpdateQuery  = "update CalendarPeriodNew set PeriodSchemaID = %s where PeriodID = %s"%(periodidpri,str(re[0]))
                #         cursor.execute(UpdateQuery)
                #         connection.commit()
                #         print("updated")
                #     except Exception as e:
                #         print(e)





    except Exception as e:
        print("Error while connecting to MySQL",e)

periods(['2000'])


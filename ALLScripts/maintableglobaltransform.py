import mysql.connector
from openpyxl import load_workbook

try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='Zeus',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
        # result = cursor.execute("select * from DWHMainQuery")
        # print(result)
        file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
        wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/CDEV_Q2 19_ V111.xlsx')
        sheet = wb.get_sheet_by_name('Industry Template')
        max_col = (sheet.max_column)
        max_row = (sheet.max_row)

except Exception as e:
    print("Error while connecting to MySQL",e)

def insertcombinations(parameters):

            try:
                result = []
                ticker = (parameters.get('Ticker'))

                #for finding coid
                for key,value in parameters.items():
                    # print( key,value)
                    if(key == 'Ticker'):
                        coidquery = "select CoID from CoSecurity where Ticker like %s"
                        cursor.execute(coidquery,(value,))
                        coid = cursor.fetchall()[0][0]
                        print(coid)
                        break
                flag = 0

                for key, value in parameters.items():
                    for row in range(1, max_row):
                        for col in range(1, max_col):
                            if(key == 'GXbrlEleID'):
                                for val in value:
                                    if val == sheet.cell(row = row,column = col).value :
                                        rowconst = row
                                        flag = 1
                                        break
                            if (flag == 1):
                                break
                        if flag == 1:
                            break
                print(rowconst)
                segmentcol = None
                productcol = None
                geocol = None
                giccol = None
                otherelementcol = None
                subelementcol = None

                for key, value in parameters.items():
                    for col in range(1, max_col):
                        if (key == 'SegmentID'):
                            for val in value:
                                if val == sheet.cell(row=rowconst, column=col).value:
                                    segmentcol = col
                                    break
                        elif(key == 'ProductID'):
                            for val in value:
                                if val == sheet.cell(row=rowconst, column=col).value:
                                    productcol = col
                                    break
                        elif (key == 'GeoID'):
                            for val in value:
                                if val == sheet.cell(row=rowconst, column=col).value:
                                    geocol = col
                                    break
                        elif (key == 'GicID'):
                            for val in value:
                                if val == sheet.cell(row=rowconst, column=col).value:
                                    giccol = col
                                    break
                        elif (key == 'OtherElementID'):
                            for val in value:
                                if val == sheet.cell(row=rowconst, column=col).value:
                                    otherelementcol = col
                                    break
                        elif (key == 'SubElementID'):
                            for val in value:
                                if val == sheet.cell(row=rowconst, column=col).value:
                                    subelementcol = col
                                    break


                print(segmentcol,productcol,geocol,giccol,otherelementcol,subelementcol)
                # rowconst = rowconst + 1
                # print(max_row)
                for rows in range(rowconst,max_row):
                    print(rows)
                    if(segmentcol == None):
                        SegmentID = None
                    else:
                        SegmentID = sheet.cell(row = rows,column = segmentcol).value

                    if(productcol == None):
                        ProductID  = None
                    else:
                        ProductID = sheet.cell(row = rows,column = productcol).value

                    if(geocol == None):
                        GeoID = None
                    else:
                        GeoID = sheet.cell(row=rows, column=geocol).value

                    if (giccol == None):
                        GicID = None
                    else:
                        GicID = sheet.cell(row=rows, column=giccol).value

                    if (otherelementcol == None):
                       OtherElementID = None
                    else:
                        OtherElementID = sheet.cell(row=rows, column=otherelementcol).value

                    if (subelementcol == None):
                        SubElementID = None
                    else:
                        SubElementID = sheet.cell(row=rows, column=subelementcol).value

                    print(SegmentID,ProductID,GeoID,GicID,OtherElementID,SubElementID)


                # flag = 0
                # for row in range(1, max_row):
                #     for col in range(1, max_col):
                #         for key, value in parameters.items():
                #             if(key == 'GXbrlEleID'):
                #                 for val in value:
                #                     if val == sheet.cell(row = row,column = col).value :
                #                         rowconst = row
                #                         flag = 1
                #                         break
                #             if(flag == 1):
                #                 break
                #         if(flag == 1):
                #             break
                # for col in range (1,max_col):
                #     for key, value in parameters.items():
                #             if (key == 'SegmentID'):
                #                 for val in value:
                #                     if val == sheet.cell(row=row, column=col).value:
                #                         segmentcol = col
                #                         break
                #             if(key == 'ProductID'):
                #                 for val in value:
                #                     if val == sheet.cell(row=row, column=col).value:
                #                         productcol = col
                #                         break
                #             # if(key == '')




                    # if ( sheet[row_ex].value == 'GXBRL ID' or  sheet[row_ex].value == 'GXbrleleID'):
                    #     rowconst = row
                #
                # for col in range(1,max_col):
                #     val = (sheet.cell(row = rowconst,column = col).value)
                #     if val = ''



            except Exception as e:
                raise e

ticker ='%' +  'CDEV'
gxbrlid = ['GXbrleleID','GXBRL ID']
segmentid = ['DivID','SegmentID']
productid = ['IsActive','ProductID']
geoid = ['Geography','IsHidden']
gicid = ['CyCalcType','GICSID']
otherelementid = ['Other ElementID','RefElementID']
subelementid =['TransformID','SubelementID']



parameters = {'Ticker':ticker,'GXbrlEleID': gxbrlid,'SegmentID':segmentid,'ProductID':productid,'GeoID':geoid,'GicID':gicid,'OtherElementID':otherelementid,'SubElementID':subelementid}
insertcombinations(parameters)

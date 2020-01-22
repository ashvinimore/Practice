import mysql.connector
from openpyxl import load_workbook
from datetime import date



def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def dumpdata(tablename):
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            cursor.execute(sqlquery)
            file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/Data Warehouse/DWH-1.xlsx',data_only=True)
            sheet = wb.get_sheet_by_name('Industry Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print(max_row,max_col)

            try:
                rowconst = ''
                descrow = ''
                flag = 0
                for col in range(1, max_col):
                    letter = (colnum_string(col))
                    for row in range(1, max_row):
                        row_ex = letter + str(row)

                        if (sheet[row_ex].value == 'GXBRL ID'):
                            rowconst = row
                            gxbrletter = letter
                            rowval = letter + str(rowconst)
                            flag = 1
                            break
                    if flag == 1:
                        # print(rowconst)
                        break
                flagsdes = 0
                rowperiod = ''
                for col in range(26,max_col):
                    for row in range(7,max_row):
                        if (sheet.cell(row=row, column=col).value  == 'Description'):
                            clnum = col + 1
                            rowperiod = row
                            flagsdes = 1
                            break
                    if flagsdes == 1:
                        break

                print("descrow",clnum)
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])
                # periodcolumn = (','.join(periodcolumn))

            #     rowseg = ''
            #     roegic = ''
            #     rowgeo = ''
            #     rowprod = ''
            #
                for col in range(10, max_col):
                    letter = (colnum_string(col))
                    # print(letter)
                    row_ex = letter + str(rowconst)

                    if (sheet[row_ex].value == 'SegmentID'):
                        rowseg = letter
                    elif (sheet[row_ex].value == 'GICSID'):
                        roegic = letter
                    elif (sheet[row_ex].value == 'Geography'):
                        rowgeo = letter
                    elif (sheet[row_ex].value == 'ProductID'):
                        rowprod = letter

                    if col >= clnum:
                        # print(clnum)
                        p = sheet.cell(row = rowperiod,column = col).value
                        p = p.replace(' ','')

                        try:
                            vperiods = str(repr(periodcolumn).replace('[','(').replace(']',')').replace("'",''))
                            periodsearchquery = "SELECT PeriodIDPri FROM CalendarPeriod WHERE %s in {}".format(vperiods)
                            cursor.execute(periodsearchquery,(p,))
                            periodid = cursor.fetchall()
                            periodid = periodid[0][0]
                        except Exception as e:
                            print(e)
                        for rowval in range(rowperiod,max_row):
                            rowvaluess =  sheet.cell(row = rowval,column = col).value
                            print(periodid,rowvaluess)
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where Date = %s"
                    cursor.execute(periodsearchquery, (today,))
                    records = cursor.fetchall()
                    print(records[0][0])
                except Exception as e:
                    print(e)



            #
            #
            #     for row in range(7, max_row):
            #         # print(row)
            #         result = []
            #         gxbrrowid =   gxbrletter + str(row)
            #         GxbrlID =  str(sheet[gxbrrowid].value)
            #         if rowseg == '':
            #             segments = 'None'
            #             segmentID = 'None'
            #         else:
            #             segments = str(rowseg) + str(row)
            #             segmentID = str(sheet[segments].value)
            #             # print(segmentID)
            #         if roegic == '':
            #             gicid = 'None'
            #             GICID = 'None'
            #         else:
            #             gicid = str(roegic) + str(row)
            #             GICID = str(sheet[gicid].value)
            #         if rowgeo == '':
            #             geoid = 'None'
            #             geography = 'None'
            #         else:
            #             geoid = str(rowgeo) + str(row)
            #             geography = str(sheet[geoid].value)
            #         if rowprod == '':
            #             prodid = 'None'
            #             productID = 'None'
            #         else:
            #             prodid = str(rowprod) + str(row)
            #             productID =  str(sheet[prodid].value)
            #
            #
            #         try :
            #             cursor.execute("SELECT UID FROM  DWHMainQuery WHERE  GXbrlEleID = %s and SegmentID = %s and ProductID = %s and GeoID = %s and GicId = %s",(GxbrlID, segmentID,productID,geography, GICID,))
            #             result = cursor.fetchall()[0][0]
            #             # print(result,GxbrlID)
            #         except Exception as e:
            #             pass




            except Exception as e:
                print(e)
                pass
    except Exception as e:
        print(e)
        pass


tablename = "`StarSchema1` (`FactID` smallint(6) NOT NULL, `UID` smallint(6) NOT NULL,  `PeriodID` varchar(50) NOT NULL,  `ReportedDateID` varchar(256) DEFAULT NULL,  `Value` int(11) DEFAULT NULL,  PRIMARY KEY (`FactID`)) ENGINE=InnoDB DEFAULT CHARSET=latin1;"
dumpdata(tablename)




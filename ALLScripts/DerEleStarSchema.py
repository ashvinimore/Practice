import mysql.connector
from openpyxl import load_workbook
from datetime import date
try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='DWHDB',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
        # result = cursor.execute("select * from DWHMainQuery")
        # print(result)

        wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW Excel/LPI_Q2 19_V1.xlsx',data_only=True)

        sheet = wb.get_sheet_by_name('Ratios')
        max_col = (sheet.max_column)
        max_row = (sheet.max_row)
        priid = 0
        periodrow = 0
        periodcol = 0

        for row in range(1, max_row):
            for col in range(1, max_col):
                if (sheet.cell(row=row, column=col).value == 'Description'):
                    periodrow = row
                    periodcol = col
        # print(periodrow,periodcol)
        #
        periodcolumn = []
        tablename = 'CalendarPeriod'
        sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
        cursor.execute(sql1)
        records = cursor.fetchall()
        for val in records:
            periodcolumn.append(val[0])
        # print(periodcolumn)

        # for todays date
        try:
            today = date.today()
            reporteddatequery = "select * from CalendarDate where YYYYMMDD = %s"
            cursor.execute(reporteddatequery, (today,))
            reporteddate = cursor.fetchall()
            reporteddate = reporteddate[0][0]
            # print(reporteddate)
        except Exception as e:
            print(e)




        coid = 9020
        periodconvert = periodcol + 1


        for rowall in range(1, max_row):
            # print("periodcolumn", periodconvert, periodrow,sheet.cell(row = periodrow,column = periodconvert).value)
            for colperiod in range(periodconvert, max_col):

                value = (sheet.cell(row=rowall, column=colperiod).value)
                # print(value)

                try:
                    if value != None and int(value) >= 0:
                        periodvalue = (sheet.cell(row=periodrow, column=colperiod).value)
                        periodvalue = periodvalue.replace(' ', '')
                        vperiods = str(repr(periodcolumn).replace('[', '(').replace(']', ')').replace("'", ''))
                        periodsearchquery = "SELECT PeriodIDPri FROM CalendarPeriod WHERE %s in {}".format(vperiods)
                        cursor.execute(periodsearchquery, (periodvalue,))
                        periodids = cursor.fetchall()
                        periodid = periodids[0][0]

                        try:
                            alreadyperiod = "select PeriodID from DWDerEleIDSchema1 where PeriodID= %s and DWCoID = %s" %(periodid,coid)
                            cursor.execute(alreadyperiod)
                            reperiod = cursor.fetchall()
                            print(reperiod[0][0])
                        except Exception as e:
                            print(e)

                            insertQuery = "INSERT INTO DWDerEleIDSchema1 (PeriodID,ReportedDateID,DWCoID) VALUES (%s,%s,%s)" %(periodid,reporteddate,coid)
                            cursor.execute(insertQuery)
                            connection.commit()
                            print("insert")

        #
                        if sheet.cell(row=rowall, column=1).value != None:
                            ratioids = str(sheet.cell(row=rowall, column=1).value)
                            valueratio = str(sheet.cell(row=rowall, column=colperiod).value)
                            # print(periodid,ratioids,valueratio)
                            try:
                                ratioupdate = "Update DWDerEleIDSchema1 set `%s` = %s where PeriodID = %s and DWCoID  = %s " %(ratioids,valueratio,periodid,coid)
                                cursor.execute(ratioupdate)
                                connection.commit()
                                print("successful")
                            except Exception as e:
                                print(e)
                                pass
                                          #
                except Exception as e:
                    # print(e)
                    pass
                else:
                    pass




except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")
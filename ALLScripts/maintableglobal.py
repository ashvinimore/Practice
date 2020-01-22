import os
import re
import csv
import mysql.connector
from openpyxl import load_workbook

try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='Zeus',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)

        wb = load_workbook('/home/pavan/Desktop/Data Warehouse/global template.xlsx')
        sheet = wb.get_sheet_by_name('Global Template')
        max_col = (sheet.max_column)
        max_row = (sheet.max_row)
        print(max_col,max_row)
        autoid =0
        try:
            result = []
            for row in range(1, max_row):
                row_ex = 'L' + str(row)
                if ( sheet[row_ex].value == 'GXBRL ID'):
                    print(row)
                    start = row + 1
                    for val in range(start,max_row):
                        rowval = 'L' + str(val)                    #
                        gxbrl = sheet[rowval].value
        #                 #
                        if(gxbrl != None ):
                            # print(gxbrl)     #        #
                            # segments,geoid, gicid, prodid = str(None)
                            cursor.execute(
                                 "SELECT GXbrlEleID FROM  DWHMainQuery WHERE  GXbrlEleID = '%s' ",
                            (gxbrl,))
                            result = (cursor.fetchone())
                            # print(result,gxbrl)
                            if(result==None):
                                autoid = autoid + 1
                                UID = autoid
                                print(UID,gxbrl)
                                parent = None
                                insertQuery = "insert into DWHMainQuery(UID,GXbrlEleID,SegmentID,GeoID,GicID,ProductID)  values(" \
                                              "'%s','%s','%s','%s','%s','%s')" % (
                                                  (int(UID),int(gxbrl),'None','None','None','None')
                                              )
                                # insertQuery = "insert into GlobalElement(GXbrlEleID,Description) values('1','1')"
                                try:
                                    cursor.execute(insertQuery)
                                    print("query excuted successfully")
                                    connection.commit()
                                except Exception as e:
                                    print(e)

                            else:
                                print("already ")

        except Exception as e:
            print(e)
            raise e

except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")
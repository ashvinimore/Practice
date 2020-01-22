import mysql.connector
from openpyxl import load_workbook
from datetime import date
from pymongo import MongoClient


#
# def colnum_string(n):
#     string = ""
#     while n > 0:
#         n, remainder = divmod(n - 1, 26)
#         string = chr(65 + remainder) + string
#     return string

def DWPeriodSchema():
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            client = MongoClient()
            client = MongoClient('localhost', 27017)
            db = client.DWHDB
            #
            # sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            # cursor.execute(sqlquery)
            # # file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW Excel/CDEV_Q2 19.xlsx',data_only=True)
            sheet = wb.get_sheet_by_name('Global Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print("max:",max_row,max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0
            othereleid = 0
            subelementid = 0
            resultpri = 355
            #
            try:
                for row in range(1,max_row):
                    for col in range(1,max_col):
                        if sheet.cell(row = row,column = col).value == 'GXBRL ID' or sheet.cell(row = row,column = col).value == 'GXbrleleID':
                          gxbrlid = col
                        if sheet.cell(row = row,column = col).value == 'SegmentID' or sheet.cell(row = row,column = col).value == 'DivID':
                            SegmentID = col
                        if sheet.cell(row = row,column = col).value == 'GICSID' or sheet.cell(row = row,column = col).value == 'CyCalcType':
                            GICSID = col
                        if sheet.cell(row = row,column = col).value == 'Geography' or sheet.cell(row = row,column = col).value == 'IsHidden':
                            Geography = col
                        if sheet.cell(row = row,column = col).value == 'ProductID' or sheet.cell(row = row,column = col).value == 'IsActive':
                            ProductID = col
                        if (sheet.cell(row=row, column=col).value == 'Description'):
                            periodrow = row
                            periodcol = col
                        if(sheet.cell(row = row,column = col).value == 'TransformID') or (sheet.cell(row = row,column = col).value == 'SubelementID') :
                            subelementid = col
                        if (sheet.cell(row=row, column=col).value == 'RefElementID') or (sheet.cell(row=row, column=col).value == 'Other ElementID'):
                            othereleid = col


            #
            #     #extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])
                print(periodcolumn)

            #     #for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where YYYYMMDD = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)            #
            #
            #




                for rowall in range(11, max_row):
                    desc = None
                    if (gxbrlid != 0 or gxbrlid != 'None'):
                        gxbrlvalue = str(sheet.cell(row = rowall,column = gxbrlid ).value)

                        try:
                            if gxbrlvalue != None :
                                if int(gxbrlvalue)  != None:
                                    globalelement = (list(db.GlobalElement.find({"GXbrlEleID":int(gxbrlvalue)})))
                                    for glob in (globalelement):
                                        desc = (glob['Description'])

                        except Exception as e:
                            print(e)
                            pass
                        if (SegmentID == 0):
                            seg = 'IS NULL'
                        else:
                            seg = str(sheet.cell(row=rowall, column=SegmentID).value)
                            if str(seg).isdigit() != True:
                                seg = 'IS NULL'
                            else:
                                seg = '= ' + str(sheet.cell(row=rowall, column=SegmentID).value)
                        if (Geography == 0):
                            dwgeoid = 'IS NULL'
                        else:
                            geographyvalue = str(sheet.cell(row=rowall, column=Geography).value)
                            if str(geographyvalue).isdigit() != True:
                                dwgeoid = 'IS NULL'
                            else:
                                geographyvalue = '=' + str (sheet.cell(row=rowall, column=Geography).value)
                                try:
                                    geoid = "select DWGeoID from DWGeography where GeoRefID  {}".format(geographyvalue)
                                    cursor.execute(geoid)
                                    dwgeoid = "= " + str(cursor.fetchall()[0][0])
                                except Exception as e:
                                    print(e)
                                    dwgeoid = 'IS NULL'
                        if (GICSID == 0):
                            dwjgicsid = 'IS NULL'
                        else:
                            GICSvalue =  str(sheet.cell(row=rowall, column=GICSID).value)
                            if str(GICSvalue).isdigit() != True:
                                dwjgicsid = 'IS NULL'
                            else:
                                GICSvalue = '=' + str(sheet.cell(row=rowall, column=GICSID).value)
                                try:
                                    gicsid = "select DWGroupID from DWGroups1 where GroupRefID {}".format(GICSvalue)
                                    cursor.execute(gicsid)
                                    dwjgicsid = "= " +  str(cursor.fetchall()[0][0])
                                except Exception as e:
                                    print(e)
                                    dwjgicsid = 'IS NULL'
                        if (ProductID == 0):
                            Productvalue =  'IS NULL'
                        else:
                            Productvalue =  str(sheet.cell(row=rowall, column=ProductID).value)
                            if str(Productvalue).isdigit() != True:
                                Productvalue = 'IS NULL'
                            else:
                                Productvalue = '= ' + str(sheet.cell(row=rowall, column=ProductID).value)
                        if (subelementid == 0):
                            subelementvalue =  'IS NULL'
                        else:
                            subelementvalue =  str(sheet.cell(row=rowall, column=subelementid).value)
                            if str(subelementvalue).isdigit() != True:
                                subelementvalue = 'IS NULL'
                            else:
                                subelementvalue = '=' + str(sheet.cell(row=rowall, column=subelementid).value)
                        if (othereleid == 0):
                            otherelevalue = 'IS NULL'
                        else:
                            otherelevalue =  str(sheet.cell(row=rowall, column=othereleid).value)
                            if str(otherelevalue).isdigit() != True:
                                otherelevalue = 'IS NULL'
                            else:
                                otherelevalue = '=' + str(sheet.cell(row=rowall, column=othereleid).value)

                        try:
                            if(gxbrlvalue.isdigit() == True):
                                gxbrlvalue = str(gxbrlvalue)
                                # print("gxbrl",gxbrlvalue,"segmentid ",seg,"geography",dwgeoid,"gicsid",dwjgicsid,"productvalue",Productvalue,"subelementvalue",subelementvalue,"otherelevalue",otherelevalue)
                                try:
                                    checkifexist = "select * from DWElement1 where EleRefID = {}".format(\
                                        str(gxbrlvalue)) + " and SegmentID {}".format(\
                                        seg) + " and ProductID {}".format(Productvalue) + " and DWGeoID {}".format(\
                                        dwgeoid) + " and DWGroupID {}".format(dwjgicsid) + " and OtherElementID {}".format(\
                                        otherelevalue) + " and SubElementID {}".format(subelementvalue)\

                                    cursor.execute(checkifexist)
                                    dwhids = cursor.fetchall()
                                    if dwhids:
                                        dwhid = dwhids[0][0]
                                        print(dwhid)
                                        if dwhid:
                                                periodconvert = 0
                                                periodconvert = periodcol + 1
                                                print("periodconvert",max_col)

                                                #
                                                for colperiod in range(periodconvert, max_col):
                                                    value = (sheet.cell(row=rowall, column=colperiod).value)
                                                    if value != None:
                                                        if int(value) != None :
                                                            # print("value", value)
                                                            coid = 11569
                                                            periodvalue = (sheet.cell(row=periodrow, column=colperiod).value)
                                                            periodvalue = periodvalue.replace(' ', '')

                                                            try:
                                                                vperiods = str(repr(periodcolumn).replace('[', '(').replace(']',
                                                                                                                            ')').replace(
                                                                    "'", ''))

                                                                periodsearchquery = "SELECT PeriodID FROM CalendarPeriod WHERE %s in {}".format(
                                                                    vperiods)
                                                                cursor.execute(periodsearchquery, (periodvalue,))
                                                                periodid = cursor.fetchall()
                                                                periodid = periodid[0][0]
                                                                print(dwhid, periodid, reporteddate,value,coid)
                                                                try:
                                                                    insertQuery = "insert into DWPeriodSchema (DWID,PeriodID,ReportedDateID,Value,DWCoID) Values(%s,%s,%s,%s,%s)"\
                                                                    %(dwhid,periodid,reporteddate,value,coid)
                                                                    cursor.execute(insertQuery)
                                                                    connection.commit()
                                                                except Exception as e:
                                                                    print("insert values", e)
                                                                    pass
                                                            except Exception as e:
                                                                print("periods",e)
                                                                pass
                                    else:
                                        print("gxbrl", gxbrlvalue, "segmentid ", seg, "geography", dwgeoid, "gicsid",
                                              dwjgicsid, "productvalue", Productvalue, "subelementvalue",
                                              subelementvalue, "otherelevalue", otherelevalue)
                                        print("not found")

                                except Exception as e:
                                    # print("outer",e)
                                    pass
                        except Exception as e:
                            pass            #
            # # #
            except Exception as e:
                print(e)

    except Exception as e:
        print(e)






DWPeriodSchema()





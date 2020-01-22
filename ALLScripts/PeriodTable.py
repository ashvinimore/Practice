# import pandas as pd
# datelist = pd.date_range(pd.datetime.today(), periods=10000).tolist()
# for val in datelist:
#    val = str(val)
#    print(val.split(' ')[0])

import mysql.connector
from openpyxl import load_workbook


import datetime
def periods(periodss):
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            record = cursor.fetchone()
            print("Your connected to - ", record)
            # file_name = '/home/pavan/Music/Data Warehouse/Period/period.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/PeriodFinal1.xlsx')
            # print(wb.get_sheet_names())

            sheet = wb.get_sheet_by_name('Period')
            max_col = (sheet.max_column)
            # print(max_col)
            max_row = (sheet.max_row) + 1
            # print(max_row)

            for row in range(5, 30):
                row_exidx = 'C' + str(row)
                row_exmodeid = 'D' + str(row)
                row_exmode = 'E' + str(row)
                row_exstart = 'J' + str(row)
                row_exend = 'K' + str(row)
                row_modepre = 'N' + str(row)

                for periodyear in periodss:
                    periodidpri = int(periodyear  + str(sheet[row_exmodeid].value)  + str(sheet[row_exidx].value))
                    periodid = int(periodyear + str(sheet[row_exidx].value) + str(sheet[row_modepre].value))
                    print(periodidpri,periodid)
                    try:
                        UpdateQuery  = "update CalendarPeriod set PeriodIDPri = %s where PeriodID = %s"%(periodidpri,periodid)
                        cursor.execute(UpdateQuery)
                        connection.commit()
                        print("updated")
                    except Exception as e:
                        print(e)


                    # periodname1 = str(sheet[row_exmode].value) +  str(sheet[row_exidx].value)  + periodyear
                    # periodname2 =  str(sheet[row_exmode].value) +  str(sheet[row_exidx].value)  +  periodyear[2:]
                    # periodname3 = str(sheet[row_exmode].value) + str(sheet[row_exidx].value) + "FY-" + periodyear
                    # periodname4 = str(sheet[row_exmode].value) + str(sheet[row_exidx].value) + "FY-" + periodyear[2:]
                    # modename =  str(sheet[row_exmode].value)
                    #
                    # if (str(sheet[row_exidx].value).startswith( '0' ) == True):
                    #     replacezero = str(sheet[row_exidx].value).replace('0', '')
                    # else:
                    #     replacezero = str(sheet[row_exidx].value)
                    # # # # print((sheet[row_exidx].value))
                    # periodname5 =  str(sheet[row_exmode].value) + str(replacezero) +  periodyear
                    # periodname6 =  str(sheet[row_exmode].value) + str(replacezero)  +  periodyear[2:]
                    # if  'YTD' in str(sheet[row_exmode].value):
                    #     # print( str(sheet[row_exmode].value))
                    #     modename2 =  str(sheet[row_exmode].value).strip('YTD')
                    #     modename2 = 'YTD' +  modename2
                    #     periodname7 =modename2 + str(sheet[row_exidx].value) + periodyear[2:]
                    #     periodname8 = modename2 + replacezero +  periodyear[2:]
                    #     periodname9 = modename2 + str(sheet[row_exidx].value)+  periodyear
                    #     periodname10 =  modename2 + replacezero +  periodyear
                    #     periodname11 =  modename2 + replacezero +  periodyear[2:]
                    # else:
                    #     modename2 =  None
                    #     periodname7 = None
                    #     periodname8 = None
                    #     periodname9 = None
                    #     periodname10 = None
                    #     periodname11 = None
                    #     periodname12 = None

                    # if  str(sheet[row_exmode].value) == 'Y':
                    #     print(str(sheet[row_exmode].value))
                    #     periodname12 = 'FY'+ periodyear[2:]
                    # else:
                    #     periodname12 = None
                    # print(modename7)
                    # print(periodidpri,periodname1,periodname2,periodname3,periodname4,periodname5,periodname6,periodname7,periodname8,periodname9,periodname10,periodname11)
                    # periodname12 = 'FY' + periodyear[2:]
                    # # To update certain values if new column added
                    # updatequery = "update CalendarPeriod set periodname12 = %s where  Mode = 'Y' and PeriodId = %s"
                    # try:
                    #     cursor.execute(updatequery,(periodname12,periodyear,))
                    #     print("query excuted successfully")
                    #     connection.commit()
                    #     # break
                    # except Exception as e:
                    #     print(e)
                    # print(periodname12)



                    # stratdate = periodyear +  str(sheet[row_exstart].value)[4:10]
                    # enddate =  periodyear +  str(sheet[row_exend].value)[4:10]
                    # startdateid = int((periodyear +  str(sheet[row_exstart].value)[4:10]).replace('-',''))
                    # enddateid =  int((periodyear +  str(sheet[row_exend].value)[4:10]).replace('-',''))
                    # # print( str(sheet[row_exidx].value))
                    # insertQuery = "INSERT INTO CalendarPeriod (PeriodIDPri,PeriodID,PeriodIDx,ModeID,Mode,PeriodName1,PeriodName2,PeriodName3,PeriodName4,StartDate,EndDate,StartDateID,EndDateID,PeriodName5,PeriodName6,ModeName2,PeriodName7,PeriodName8,PeriodName9,PeriodName10,PeriodName11) VALUES (""'%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')" %(periodidpri,periodyear,(sheet[row_exidx].value),str(sheet[row_exmodeid].value),modename,periodname1,periodname2,periodname3,periodname4,stratdate,enddate,startdateid,enddateid,periodname5,periodname6,modename2,periodname7,periodname8,periodname9,periodname10,periodname11,)
                    #
                    # try:
                    #     cursor.execute(insertQuery)
                    #     print("query excuted successfully")
                    #     connection.commit()
                    # except Exception as e:
                    #     print(e)

        # start = datetime.datetime.strptime("01-01-1995", "%d-%m-%Y")
        # end = datetime.datetime.strptime("31-12-2050", "%d-%m-%Y")
        # date_generated = [start + datetime.timedelta(days=x) for x in range(0, (end-start).days)]
        #
        # for date in date_generated:
        #     yydate =  (date.strftime("%Y-%m-%d"))
        #     standarddate = str (date.strftime("%d-%m-%Y"))
        #     dateid =  (date.strftime("%Y%m%d"))
        #     print(dateid)
        #     insertQuery = "insert into CalendarDate(DateID,Date,StandardDateFormat)  values(""'%s','%s','%s')" % ((dateid, yydate, standarddate,))
        #     try:
        #         cursor.execute(insertQuery)
        #         print("query excuted successfully")
        #         connection.commit()
        #     except Exception as e:
        #         print(e)



    except Exception as e:
        print("Error while connecting to MySQL",e)

periods(['2051','2052','2053','2054','2055','2056','2057','2058','2059','2060'])


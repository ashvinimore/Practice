import mysql.connector
from openpyxl import load_workbook
from datetime import date
import pandas as pd


def DailyPeriods(Periods):
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            record = cursor.fetchone()
            print("Your connected to - ", record)
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DailyPeriod1.xlsx',data_only=True)
            sheet = wb.get_sheet_by_name('Sheet2')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print(max_col,max_row)

            for period in Periods:
                for row in range(2,max_row):
                    year = period
                    day =  sheet.cell(row = row,column = 2).value
                    date = sheet.cell(row = row,column = 5).value
                    month =  sheet.cell(row = row,column = 6).value
                    periodid = str(year) + '06'+str(day)
                    periodname =str(date) + '/' + str(month)+ '/' + str(year)
                    print(year, day, date, month,periodid,periodname)
                    insertQuery = "INSERT INTO CalendarDailyPeriod (PeriodID,PeriodYear,PeriodMonth,PeriodIDx,ModeID,Mode,PeriodName1) values (%s,%s,%s,%s,%s,'%s','%s')" \
                                  %(periodid,str(period),str(month),str(day),'06',str('D'),str(periodname))
                    cursor.execute(insertQuery)
                    connection.commit()
                    # break
                    print("inserted")


    except Exception as e:
        print("Error while connecting to MySQL",e)
        pass

    finally:
        # closing database connection.
        if (connection.is_connected()):
            cursor.close()
            connection.close()
            print("MySQL connection is closed")

Periods = [2060]
DailyPeriods(Periods)
import mysql.connector
from openpyxl import load_workbook
from datetime import date
from pymongo import MongoClient


#
# def colnum_string(n):
#     string = ""
#     while n > 0:
#         n, remainder = divmod(n - 1, 26)
#         string = chr(65 + remainder) + string
#     return string

def dumpdataStarSchema():
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            client = MongoClient()
            client = MongoClient('localhost', 27017)
            db = client.DWHDB
            #
            # sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            # cursor.execute(sqlquery)
            # # file_name = '/home/pavan/Music/Data Warehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW Excel/CDEV_Q2 19.xlsxx',data_only=True)
            sheet = wb.get_sheet_by_name('Industry Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print("max:",max_row,max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0
            othereleid = 0
            subelementid = 0
            #
            try:
                for row in range(1,max_row):
                    for col in range(1,max_col):
                        if sheet.cell(row = row,column = col).value == 'GXBRL ID' or sheet.cell(row = row,column = col).value == 'GXbrleleID':
                          gxbrlid = col
                        if sheet.cell(row = row,column = col).value == 'SegmentID' or sheet.cell(row = row,column = col).value == 'DivID':
                            SegmentID = col
                        if sheet.cell(row = row,column = col).value == 'GICSID' or sheet.cell(row = row,column = col).value == 'CyCalcType':
                            GICSID = col
                        if sheet.cell(row = row,column = col).value == 'Geography' or sheet.cell(row = row,column = col).value == 'IsHidden':
                            Geography = col
                        if sheet.cell(row = row,column = col).value == 'ProductID' or sheet.cell(row = row,column = col).value == 'IsActive':
                            ProductID = col
                        if (sheet.cell(row=row, column=col).value == 'Description'):
                            periodrow = row
                            periodcol = col
                        if(sheet.cell(row = row,column = col).value == 'TransformID') or (sheet.cell(row = row,column = col).value == 'SubelementID') :
                            subelementid = col
                        if (sheet.cell(row=row, column=col).value == 'RefElementID') or (sheet.cell(row=row, column=col).value == 'Other ElementID'):
                            othereleid = col


            #
            #     #extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])
                print(periodcolumn)

            #     #for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where YYYYMMDD = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)            #
            #
            #


                for rowall in range(1, max_row):
                    desc = None
                    if (gxbrlid != 0 or gxbrlid != 'None'):
                        gxbrlvalue = str(sheet.cell(row = rowall,column = gxbrlid ).value)
                        # print(gxbrlvalue)
                        try:
                            if gxbrlvalue != 'None' :
                                if int(gxbrlvalue)  != None:
                                    globalelement = (list(db.GlobalElement.find({"GXbrlEleID":int(gxbrlvalue)})))
                                    for glob in (globalelement):
                                        desc = (glob['Description'])

                        except Exception as e:
                            print(e)
                            pass


                        if (SegmentID == 0):
                            seg = 'IS NULL'
                        else:
                            seg = str(sheet.cell(row=rowall, column=SegmentID).value)
                            if str(seg).isdigit() != True:
                                seg = 'IS NULL'
                            else:
                                seg = '=' + str(sheet.cell(row=rowall, column=SegmentID).value)
                        if (Geography == 0):
                            dwgeoid = 'IS NULL'
                        else:
                            geographyvalue = str(sheet.cell(row=rowall, column=Geography).value)
                            if str(geographyvalue).isdigit() != True:
                                dwgeoid = 'IS NULL'
                            else:
                                geographyvalue = '=' + str (sheet.cell(row=rowall, column=Geography).value)
                                try:
                                    geoid = "select DWGeoID from DWGeography where GeoRefID  {}".format(geographyvalue)
                                    cursor.execute(geoid)
                                    dwgeoid = "= " + str(cursor.fetchall()[0][0])
                                except Exception as e:
                                    print(e)
                                    dwgeoid = 'IS NULL'


                        if (GICSID == 0):
                            dwjgicsid = 'IS NULL'

                        else:
                            GICSvalue =  str(sheet.cell(row=rowall, column=GICSID).value)
                            if str(GICSvalue).isdigit() != True:
                                dwjgicsid = 'IS NULL'
                            else:
                                GICSvalue = '=' + str(sheet.cell(row=rowall, column=GICSID).value)
                                try:
                                    gicsid = "select DWGroupID from DWGroups1 where GroupRefID {}".format(GICSvalue)
                                    cursor.execute(gicsid)
                                    dwjgicsid = "= " +  str(cursor.fetchall()[0][0])
                                except Exception as e:
                                    print(e)
                                    dwjgicsid = 'IS NULL'
                        if (ProductID == 0):
                            Productvalue =  'IS NULL'
                        else:
                            Productvalue =  str(sheet.cell(row=rowall, column=ProductID).value)
                            if str(Productvalue).isdigit() != True:
                                Productvalue = 'IS NULL'
                            else:
                                Productvalue = '=' + str(sheet.cell(row=rowall, column=ProductID).value)
                        if (subelementid == 0):
                            subelementvalue =  'IS NULL'
                        else:
                            subelementvalue =  str(sheet.cell(row=rowall, column=subelementid).value)
                            if str(subelementvalue).isdigit() != True:
                                subelementvalue = 'IS NULL'
                            else:
                                subelementvalue = '=' + str(sheet.cell(row=rowall, column=subelementid).value)
                        if (othereleid == 0):
                            otherelevalue = 'IS NULL'
                        else:
                            otherelevalue =  str(sheet.cell(row=rowall, column=othereleid).value)
                            if str(otherelevalue).isdigit() != True:
                                otherelevalue = 'IS NULL'
                            else:
                                otherelevalue = '=' + str(sheet.cell(row=rowall, column=othereleid).value)

                        try:
                            if(gxbrlvalue.isdigit() == True):
                                # print("gxbrl: ",desc,"segmentid: ",seg,"geography:",dwgeoid,"gicsid : ",dwjgicsid,"productvalue: ",Productvalue,"subelementvalue: ",subelementvalue,"otherelevalue :",otherelevalue)
                                try:
                                    checkifexist = "select * from DWElement1 where EleRefID = %s and SegmentID {}".format(\
                                        seg) + " and ProductID {}".format(Productvalue) + " and DWGeoID {}".format(\
                                        dwgeoid) + " and DWGroupID {}".format(dwjgicsid) + " and OtherElementID {}".format(\
                                        otherelevalue) + " and SubElementID {}".format(subelementvalue)\


                                    cursor.execute(checkifexist,(gxbrlvalue,))
                                    dwhids = cursor.fetchall()
                                    if dwhids:
                                        dwhid = dwhids[0][0]
                                        print(dwhid)
            #                             if  dwhid :
            #                                 periodconvert = 0
            #                                 periodconvert = periodcol + 1
            #                                 print("periodconvert",max_col)
            #                                 #
            #                                 for colperiod in range(periodconvert, max_col):
            #                                     value = (sheet.cell(row=rowall, column=colperiod).value)
            #                                     if value != None:
            #                                         if int(value) != None :
            #                                             # print("value", value)
            #                                             coid = 96
            #                                             periodvalue = (sheet.cell(row=periodrow, column=colperiod).value)
            #                                             periodvalue = periodvalue.replace(' ', '')
            #
            #                                             try:
            #                                                 vperiods = str(repr(periodcolumn).replace('[', '(').replace(']',
            #                                                                                                             ')').replace(
            #                                                     "'", ''))
            #
            #                                                 periodsearchquery = "SELECT PeriodID FROM CalendarPeriod WHERE %s in {}".format(
            #                                                     vperiods)
            #                                                 cursor.execute(periodsearchquery, (periodvalue,))
            #                                                 periodid = cursor.fetchall()
            #                                                 periodid = periodid[0][0]
            #                                                 # print(dwhid, periodid, reporteddate,value,coid)
            # #                             #
            #                                                 if periodid:
            #                                                     try:
            #                                                         searchQuery = "select DWID from DWStarSchema1 where DWID = %s and PeriodID = %s and ReportedDateID = %s and DWCoID = %s"
            #                                                         cursor.execute(searchQuery,
            #                                                                        (dwhid, periodid, reporteddate,coid,))
            #                                                         serachdwhid = cursor.fetchall()
            #                                                         if serachdwhid:
            #                                                             print("serachdwhid",serachdwhid)
            #                                                         else:
            #
            #                                                                 insertQuery = "insert into DWStarSchema1(DWID,PeriodID,ReportedDateID,Value,DWCoID) values(" \
            #                                                                               "'%s','%s','%s','%s','%s')" % (
            #                                                                                   (dwhid, periodid, reporteddate,
            #                                                                                    value,coid))
            #                                                                 try:
            #                                                                     cursor.execute(insertQuery)
            #                                                                     print("query excuted successfully")
            #                                                                     connection.commit()
            #                                                                 except Exception as e:
            #                                                                     print("insertquery", e)
            #                                                     except Exception as e:
            #                                                         print("serachquery", e)
            #
            #                                             except Exception as e:
            #                                                 print(e, periodvalue)
            #                                                 pass
            #                                     else:
            #                                         pass
                                    else:
                                        print("Result")
                                except Exception as e:
                                    pass
                        except Exception as e:
                            print(e)
                            pass            #
            # #
            except Exception as e:
                print(e)

    except Exception as e:
        print(e)



def dumpPeriodSchema():
    # print("dump uid based schema",tablename)
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            # sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            # cursor.execute(sqlquery)
            file_name = '/home/pavan/Music/DataWarehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW Excel/_NOV _Q2_191 DC.xlsx',data_only=True)
            sheet = wb.get_sheet_by_name('Global Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print(max_row,max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0

            try:
                for row in range(1, max_row):
                    for col in range(1, max_col):
                        if sheet.cell(row=row, column=col).value == 'GXBRL ID' or sheet.cell(row=row,
                                                                                             column=col).value == 'GXbrleleID':
                            gxbrlid = col
                        if sheet.cell(row=row, column=col).value == 'SegmentID' or sheet.cell(row=row,
                                                                                              column=col).value == 'DivID':
                            SegmentID = col
                        if sheet.cell(row=row, column=col).value == 'GICSID' or sheet.cell(row=row,
                                                                                           column=col).value == 'CyCalcType':
                            GICSID = col
                        if sheet.cell(row=row, column=col).value == 'Geography' or sheet.cell(row=row,
                                                                                              column=col).value == 'IsHidden':
                            Geography = col
                        if sheet.cell(row=row, column=col).value == 'ProductID' or sheet.cell(row=row,
                                                                                              column=col).value == 'IsActive':
                            ProductID = col
                        if (sheet.cell(row=row, column=col).value == 'Description'):
                            periodrow = row
                            periodcol = col
                        if (sheet.cell(row=row, column=col).value == 'TransformID') or (
                                sheet.cell(row=row, column=col).value == 'SubelementID'):
                            subelementid = col
                        if (sheet.cell(row=row, column=col).value == 'RefElementID') or (
                                sheet.cell(row=row, column=col).value == 'Other ElementID'):
                            othereleid = col

                # print(gxbrlid, SegmentID, GICSID, Geography, ProductID, Description)

                # extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])
                print(periodcolumn)

                #     #for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where Date = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)

                for rowall in range(1, max_row):
                        # print("rowall",rowall)

                        if (gxbrlid != 0):
                            gxbrlvalue = str(sheet.cell(row=rowall, column=gxbrlid).value)
                            if (SegmentID == 0):
                                seg = 'IS NULL'
                            else:
                                seg = str(sheet.cell(row=rowall, column=SegmentID).value)
                                if str(seg).isdigit() != True:
                                    seg = 'IS NULL'
                                else:
                                    seg = '=' + str(sheet.cell(row=rowall, column=SegmentID).value)
                            if (Geography == 0):
                                geographyvalue = 'IS NULL'
                            else:
                                geographyvalue = str(sheet.cell(row=rowall, column=Geography).value)
                                if str(geographyvalue).isdigit() != True:
                                    geographyvalue = 'IS NULL'
                                else:
                                    geographyvalue = '=' + str(sheet.cell(row=rowall, column=Geography).value)
                            if (GICSID == 0):
                                GICSvalue = 'IS NULL'
                            else:
                                GICSvalue = str(sheet.cell(row=rowall, column=GICSID).value)
                                if str(GICSvalue).isdigit() != True:
                                    GICSvalue = 'IS NULL'
                                else:
                                    GICSvalue = '=' + str(sheet.cell(row=rowall, column=GICSID).value)
                            if (ProductID == 0):
                                Productvalue = 'IS NULL'
                            else:
                                Productvalue = str(sheet.cell(row=rowall, column=ProductID).value)
                                if str(Productvalue).isdigit() != True:
                                    Productvalue = 'IS NULL'
                                else:
                                    Productvalue = '=' + str(sheet.cell(row=rowall, column=ProductID).value)
                            if (subelementid == 0):
                                subelementvalue = 'IS NULL'
                            else:
                                subelementvalue = str(sheet.cell(row=rowall, column=subelementid).value)
                                if str(subelementvalue).isdigit() != True:
                                    subelementvalue = 'IS NULL'
                                else:
                                    subelementvalue = '=' + str(sheet.cell(row=rowall, column=subelementid).value)
                            if (othereleid == 0):
                                otherelevalue = 'IS NULL'
                            else:
                                otherelevalue = str(sheet.cell(row=rowall, column=othereleid).value)
                                if str(otherelevalue).isdigit() != True:
                                    otherelevalue = 'IS NULL'
                                else:
                                    otherelevalue = '=' + str(sheet.cell(row=rowall, column=othereleid).value)
                            # if (gxbrlvalue .isdigit() == True):
                            # print("gxbrlvalue",gxbrlvalue,"seg",seg,"geographyvalue",geographyvalue,"GICSvalue",GICSvalue,"Productvalue",Productvalue,"subelementvalue",subelementvalue,"otherelevalue",otherelevalue)
                            try:
                                if (gxbrlvalue.isdigit() == True):
                                    # print(gxbrlvalue,seg)
                                    try:
                                        checkifexist = "select * from DWMainElement where RefID = %s and SegmentID {}".format( \
                                            seg) + " and ProductID {}".format(Productvalue) + " and GeoID {}".format( \
                                            geographyvalue) + " and GicID {}".format(
                                            GICSvalue) + " and OtherElementID {}".format( \
                                            otherelevalue) + " and SubElementID {}".format(subelementvalue) \
                                            # checkifexist = "select * from DWHMainQuery where GXbrlEleID = %s and SegmentID {}".format( \
                                        #         seg)

                                        cursor.execute(checkifexist, (gxbrlvalue,))
                                        dwhids = cursor.fetchall()

                                        coid = 115669
                                        if dwhids:
                                                dwhid = dwhids[0][0]
                                                print(dwhid)
                                                findQuery =  "select DWID from DWPeriodSchema where DWID = %s and DWCoID = %s"%(dwhid,coid)
                                                cursor.execute(findQuery)
                                                try:
                                                    result = cursor.fetchall()[0][0]
                                                    print(result)
                                                    periodconvert = 1
                                                    periodconvert = periodcol + 1
                                                    for colperiod in range(periodconvert, max_col):
                                                        value = (sheet.cell(row=rowall, column=colperiod).value)
                                                        if value != None:
                                                            # print(value)
                                                            if int(value) != None:
                                                                print("value", value)
                                                                periodvalue = (
                                                                    sheet.cell(row=periodrow, column=colperiod).value)
                                                                periodvalue = periodvalue.replace(' ', '')

                                                                try:
                                                                    vperiods = str(
                                                                        repr(periodcolumn).replace('[', '(').replace(']',
                                                                                                                     ')').replace(
                                                                            "'", ''))

                                                                    periodsearchquery = "SELECT PeriodID FROM CalendarPeriod WHERE %s in {}".format(
                                                                        vperiods)
                                                                    cursor.execute(periodsearchquery, (periodvalue,))
                                                                    periodid = cursor.fetchall()
                                                                    periodid = periodid[0][0]
                                                                    print(dwhid, periodid, reporteddate, value)
                                                                    searchDWHID = "Select DWID from DWPeriodSchema where DWID = %s"%dwhid
                                                                    cursor.execute(searchDWHID)
                                                                    try:
                                                                        se = cursor.fetchall()[0][0]
                                                                        print("sEEE",se)
                                                                        updateQuery = "Update DWPeriodSchema set `%s` = %s where DWCoID = %s and ReportedDateID = %s and DWID = %s"

                                                                        try:
                                                                            # cursor.execute(insertQuery)
                                                                            cursor.execute(updateQuery,(periodid,value,coid,reporteddate,se))
                                                                            connection.commit()
                                                                            print("updated")
                                                                        except Exception as e:
                                                                            print("e",e)
                                                                    except Exception as e:
                                                                            print("ee",e)
                                                                except Exception as e:
                                                                    print("eee",e)
                                                except Exception  as e:
                                                    # insertQuery = "Insert into DWPeriodWiseSchema1 (DWHID,ReportedDateID,DWHCoID) values  ('%s','%s','%s')" % (
                                                    # dwhid, reporteddate, coid)
                                                    # cursor.execute(insertQuery)
                                                    # connection.commit()
                                                    # print("inserted successfully")
                                                    print(e)

                                            # # if (dwhid == 12):
                                            #     periodconvert = 0
                                            #     periodconvert = periodcol + 1
                                            #     # print("periodconvert", max_col)
                                            #     #
                                            #     for colperiod in range(periodconvert, max_col):
                                            #         value = (sheet.cell(row=rowall, column=colperiod).value)
                                            #         if value != None:
                                            #             # print(value)
                                            #             if int(value) != None:
                                            #                 # print("value", value)
                                            #                 periodvalue = (
                                            #                     sheet.cell(row=periodrow, column=colperiod).value)
                                            #                 periodvalue = periodvalue.replace(' ', '')
                                            #
                                            #                 try:
                                            #                     vperiods = str(
                                            #                         repr(periodcolumn).replace('[', '(').replace(']',
                                            #                                                                      ')').replace(
                                            #                             "'", ''))
                                            #
                                            #                     periodsearchquery = "SELECT PeriodIDPri FROM CalendarPeriod WHERE %s in {}".format(
                                            #                         vperiods)
                                            #                     cursor.execute(periodsearchquery, (periodvalue,))
                                            #                     periodid = cursor.fetchall()
                                            #                     periodid = periodid[0][0]
                                            #                     print(dwhid, periodid, reporteddate, value)
                                            #                     searchDWHID = "Select DWHID from PeriodWiseSchema where DWHID = %s"%dwhid
                                            #                     cursor.execute(searchDWHID)
                                            #                     try:
                                            #                         se = cursor.fetchall()[0][0]
                                            #                         print(se)
                                            #                         updateQuery = "Update DWPeriodWiseSchema set DWHID = %s ,ReportedDateID = %s ,`%s` = %s where DWHID = %s"
                                            #                         try:
                                            #                             # cursor.execute(insertQuery)
                                            #                             cursor.execute(updateQuery,(se,reporteddate,periodid,value,se))
                                            #                             connection.commit()
                                            #                             print("executed")
                                            #                         except Exception as e:
                                            #                             print(e)
                                            #                     except Exception as e:
                                            #                         print(e)
                                                                    # insertQuery  =  "insert into PeriodWiseSchema(DWHID,ReportedDateID,`2015011`) values(" \
                                                                    #                           "'%s','%s','%s')" % (
                                                                    #                               (dwhid,  reporteddate,
                                                                    #                                value))
                                                                    # updateQuery = "Update PeriodWiseSchema set DWHID = %s ,ReportedDateID = %s ,`%s` = %s where DWHID = 12 "
                                                                    # try:
                                                                    #     # cursor.execute(insertQuery)
                                                                    #     cursor.execute(updateQuery,(dwhid, reporteddate,periodid,value))
                                                                    #     connection.commit()

                                                                    #
                                                                    #
                                                                    # except Exception as e:
                                                                    #     print(e)
                                                            # except Exception as e:
                                                            #     print(e)

                                    except Exception as e:
                                        print(e)

                            except Exception as e:
                                print(e)

            except Exception as e:
                print(e)




            # addperiodcolumnquery = "alter table  `PeriodwiseSchema1` add column `PeriodId`   smallint(6) "
    except Exception as e:
                print(e)



def dumpUIDSchema():
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            # sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            # cursor.execute(sqlquery)
            file_name = '/home/pavan/Music/DataWarehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW Excel/LPI_Q2 19_V1.xlsx',
                               data_only=True)
            sheet = wb.get_sheet_by_name('Global Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print(max_row, max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0

            try:
                for row in range(1, max_row):
                    for col in range(1, max_col):
                        if sheet.cell(row=row, column=col).value == 'GXBRL ID' or sheet.cell(row=row,
                                                                                             column=col).value == 'GXbrleleID':
                            gxbrlid = col
                        if sheet.cell(row=row, column=col).value == 'SegmentID' or sheet.cell(row=row,
                                                                                              column=col).value == 'DivID':
                            SegmentID = col
                        if sheet.cell(row=row, column=col).value == 'GICSID' or sheet.cell(row=row,
                                                                                           column=col).value == 'CyCalcType':
                            GICSID = col
                        if sheet.cell(row=row, column=col).value == 'Geography' or sheet.cell(row=row,
                                                                                              column=col).value == 'IsHidden':
                            Geography = col
                        if sheet.cell(row=row, column=col).value == 'ProductID' or sheet.cell(row=row,
                                                                                              column=col).value == 'IsActive':
                            ProductID = col
                        if (sheet.cell(row=row, column=col).value == 'Description'):
                            periodrow = row
                            periodcol = col
                        if (sheet.cell(row=row, column=col).value == 'TransformID') or (
                                sheet.cell(row=row, column=col).value == 'SubelementID'):
                            subelementid = col
                        if (sheet.cell(row=row, column=col).value == 'RefElementID') or (
                                sheet.cell(row=row, column=col).value == 'Other ElementID'):
                            othereleid = col

                # print(gxbrlid, SegmentID, GICSID, Geography, ProductID, Description)

                # extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])
                # print(periodcolumn)

                #     #for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where YYYYMMDD = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)

                for rowall in range(1, max_row):
                    # print("rowall",rowall)

                    if (gxbrlid != 0):
                        gxbrlvalue = str(sheet.cell(row=rowall, column=gxbrlid).value)
                        if (SegmentID == 0):
                            seg = 'IS NULL'
                        else:
                            seg = str(sheet.cell(row=rowall, column=SegmentID).value)
                            if str(seg).isdigit() != True:
                                seg = 'IS NULL'
                            else:
                                seg = '=' + str(sheet.cell(row=rowall, column=SegmentID).value)
                        if (Geography == 0):
                            geographyvalue = 'IS NULL'
                        else:
                            geographyvalue = str(sheet.cell(row=rowall, column=Geography).value)
                            if str(geographyvalue).isdigit() != True:
                                geographyvalue = 'IS NULL'
                            else:
                                geographyvalue = '=' + str(sheet.cell(row=rowall, column=Geography).value)
                        if (GICSID == 0):
                            GICSvalue = 'IS NULL'
                        else:
                            GICSvalue = str(sheet.cell(row=rowall, column=GICSID).value)
                            if str(GICSvalue).isdigit() != True:
                                GICSvalue = 'IS NULL'
                            else:
                                GICSvalue = '=' + str(sheet.cell(row=rowall, column=GICSID).value)
                        if (ProductID == 0):
                            Productvalue = 'IS NULL'
                        else:
                            Productvalue = str(sheet.cell(row=rowall, column=ProductID).value)
                            if str(Productvalue).isdigit() != True:
                                Productvalue = 'IS NULL'
                            else:
                                Productvalue = '=' + str(sheet.cell(row=rowall, column=ProductID).value)
                        if (subelementid == 0):
                            subelementvalue = 'IS NULL'
                        else:
                            subelementvalue = str(sheet.cell(row=rowall, column=subelementid).value)
                            if str(subelementvalue).isdigit() != True:
                                subelementvalue = 'IS NULL'
                            else:
                                subelementvalue = '=' + str(sheet.cell(row=rowall, column=subelementid).value)
                        if (othereleid == 0):
                            otherelevalue = 'IS NULL'
                        else:
                            otherelevalue = str(sheet.cell(row=rowall, column=othereleid).value)
                            if str(otherelevalue).isdigit() != True:
                                otherelevalue = 'IS NULL'
                            else:
                                otherelevalue = '=' + str(sheet.cell(row=rowall, column=othereleid).value)
                        # if (gxbrlvalue .isdigit() == True):
                        # print("gxbrlvalue",gxbrlvalue,"seg",seg,"geographyvalue",geographyvalue,"GICSvalue",GICSvalue,"Productvalue",Productvalue,"subelementvalue",subelementvalue,"otherelevalue",otherelevalue)
                        try:
                            if (gxbrlvalue.isdigit() == True):
                                # print(gxbrlvalue,seg)
                                try:
                                    checkifexist = "select * from DWElement where EleRefID= %s and SegmentID {}".format( \
                                        seg) + " and ProductID {}".format(Productvalue) + " and DWGeoID {}".format( \
                                        geographyvalue) + " and DWGroupID {}".format(
                                        GICSvalue) + " and OtherElementID {}".format( \
                                        otherelevalue) + " and SubElementID {}".format(subelementvalue) \
                                        # checkifexist = "select * from DWHMainQuery where GXbrlEleID = %s and SegmentID {}".format( \
                                    #         seg)

                                    cursor.execute(checkifexist, (gxbrlvalue,))
                                    dwhids = cursor.fetchall()
                                    coid = 9020
                                    print(dwhids)
                                    if dwhids:
                                        dwhid = 'ID' + str(dwhids[0][0])
                                        # if (dwhid == 12):
                                        periodconvert = 0
                                        periodconvert = periodcol + 1
                                        #
                                        for colperiod in range(periodconvert, max_col):
                                            value = (sheet.cell(row=rowall, column=colperiod).value)
                                            if value != None:
                                                # print(value)
                                                if int(value) != None:
                                                    # print("value", value)
                                                    periodvalue = (
                                                        sheet.cell(row=periodrow, column=colperiod).value)
                                                    periodvalue = periodvalue.replace(' ', '')

                                                    try:
                                                        vperiods = str(
                                                            repr(periodcolumn).replace('[', '(').replace(']',
                                                                                                         ')').replace(
                                                                "'", ''))

                                                        periodsearchquery = "SELECT PeriodID FROM CalendarPeriodNew WHERE %s in {}".format(
                                                            vperiods)
                                                        cursor.execute(periodsearchquery, (periodvalue,))
                                                        periodid = cursor.fetchall()
                                                        periodid = periodid[0][0]
                                                        # print(dwhid, periodid, reporteddate, value)
                                                        searchQuery = "select * from DWEleIDSchemaNew where PeriodID = %s and DWCoID = %s" % (
                                                        periodid, coid)
                                                        try:
                                                            cursor.execute(searchQuery)
                                                            re = cursor.fetchall()
                                                            # print(re)
                                                            if re == []:
                                                                print(re)
                                                                insertQuery = "insert into DWEleIDSchemaNew (PeriodID,ReportedDateID,DWCoID) values(%s,%s,%s)" % (
                                                                    periodid, reporteddate, coid,)
                                                                try:
                                                                    cursor.execute(insertQuery)
                                                                    connection.commit()
                                                                    print("query executed")
                                                                except Exception as e:
                                                                    print("error", e)

                                                            else:
                                                                print(re)
                                                                updateQuery = "Update DWEleIDSchemaNew set `%s` = %s where PeriodID = %s and ReportedDateID = %s and DWCoID = %s " % (
                                                                    str(dwhid), value, periodid, reporteddate, coid)
                                                                try:
                                                                    cursor.execute(updateQuery)
                                                                    print("executed")
                                                                    connection.commit()
                                                                except Exception as e:
                                                                    print(e)
                                                                    # print(e)


                                                        except Exception as e:
                                                            print("error", e)

                                                        # insertQuery  = "insert into DWEleIDSchema1 (PeriodID,ReportedDateID,DWCoID) values(%s,%s,%s)"%(periodid,reporteddate,coid,)
                                                        # try:
                                                        #     cursor.execute(insertQuery)
                                                        # except Exception as e:
                                                        #     print("error",e)

                                                        # searchDWHID = "Select periodid from DWEleIDSchema1 where periodid = %s and DWCoID = %s" % (periodid,coid)
                                                        # cursor.execute(searchDWHID)
                                                        # try:
                                                        #     se = cursor.fetchall()[0][0]
                                                        #     print(se)
                                                        # updateQuery = "Update DWEleIDSchema set `%s` = %s where PeriodID = %s and ReportedDateID = %s and DWCoID = %s "
                                                        # try:
                                                        #     cursor.execute(updateQuery, (
                                                        #     dwhid, value,periodid,reporteddate, coid))
                                                        #     print("executed")
                                                        #     connection.commit()
                                                        # except Exception as e:
                                                        #     print(e)
                                                        # except Exception as e:
                                                        #     print("error",e)
                                                    #
                                                    except Exception as e:
                                                        print(e)
                                #
                                except Exception as e:
                                    print(e)
                        #
                        except Exception as e:
                            print(e)

            except Exception as e:
                print(e)
    except Exception as e:
        print(e)


def dumpPeriodSchema():
    try:
        connection = mysql.connector.connect(host='192.168.10.13',
                                             database='DWHDB',
                                             user='root',
                                             password='Bluecrest')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL database... MySQL Server version on ", db_Info)
            cursor = connection.cursor(buffered=True)
            cursor.execute("select database();")
            print((cursor.fetchall()[0][0]))

            # sqlquery = "CREATE TABLE IF NOT EXISTS %s"%tablename
            # cursor.execute(sqlquery)
            file_name = '/home/pavan/Music/DataWarehouse/DWH-1.xlsx'
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW Excel/EOG_Q2 19.xlsx', data_only=True)
            sheet = wb.get_sheet_by_name('Global Template')
            max_col = (sheet.max_column)
            max_row = (sheet.max_row) + 1
            print(max_row, max_col)
            gxbrlid = 0
            SegmentID = 0
            GICSID = 0
            Geography = 0
            ProductID = 0

            try:
                for row in range(1, max_row):
                    for col in range(1, max_col):
                        if sheet.cell(row=row, column=col).value == 'GXBRL ID' or sheet.cell(row=row,
                                                                                             column=col).value == 'GXbrleleID':
                            gxbrlid = col
                        if sheet.cell(row=row, column=col).value == 'SegmentID' or sheet.cell(row=row,
                                                                                              column=col).value == 'DivID':
                            SegmentID = col
                        if sheet.cell(row=row, column=col).value == 'GICSID' or sheet.cell(row=row,
                                                                                           column=col).value == 'CyCalcType':
                            GICSID = col
                        if sheet.cell(row=row, column=col).value == 'Geography' or sheet.cell(row=row,
                                                                                              column=col).value == 'IsHidden':
                            Geography = col
                        if sheet.cell(row=row, column=col).value == 'ProductID' or sheet.cell(row=row,
                                                                                              column=col).value == 'IsActive':
                            ProductID = col
                        if (sheet.cell(row=row, column=col).value == 'Description'):
                            periodrow = row
                            periodcol = col
                        if (sheet.cell(row=row, column=col).value == 'TransformID') or (
                                sheet.cell(row=row, column=col).value == 'SubelementID'):
                            subelementid = col
                        if (sheet.cell(row=row, column=col).value == 'RefElementID') or (
                                sheet.cell(row=row, column=col).value == 'Other ElementID'):
                            othereleid = col

                # print(gxbrlid, SegmentID, GICSID, Geography, ProductID, Description)

                # extract all the columns from CalendarPeriod to check for periodname
                periodcolumn = []
                tablename = 'CalendarPeriod'
                sql1 = "SELECT  COLUMN_NAME FROM  INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = N'%s'" % tablename;
                cursor.execute(sql1)
                records = cursor.fetchall()
                for val in records:
                    periodcolumn.append(val[0])
                # print(periodcolumn)

                #     #for todays date
                try:
                    today = date.today()
                    reporteddatequery = "select * from CalendarDate where YYYYMMDD = %s"
                    cursor.execute(reporteddatequery, (today,))
                    reporteddate = cursor.fetchall()
                    reporteddate = reporteddate[0][0]
                    print(reporteddate)
                except Exception as e:
                    print(e)

                for rowall in range(1, max_row):
                    # print("rowall",rowall)

                    if (gxbrlid != 0):
                        gxbrlvalue = str(sheet.cell(row=rowall, column=gxbrlid).value)
                        if (SegmentID == 0):
                            seg = 'IS NULL'
                        else:
                            seg = str(sheet.cell(row=rowall, column=SegmentID).value)
                            if str(seg).isdigit() != True:
                                seg = 'IS NULL'
                            else:
                                seg = '=' + str(sheet.cell(row=rowall, column=SegmentID).value)
                        if (Geography == 0):
                            geographyvalue = 'IS NULL'
                        else:
                            geographyvalue = str(sheet.cell(row=rowall, column=Geography).value)
                            if str(geographyvalue).isdigit() != True:
                                geographyvalue = 'IS NULL'
                            else:
                                geographyvalue = '=' + str(sheet.cell(row=rowall, column=Geography).value)
                        if (GICSID == 0):
                            GICSvalue = 'IS NULL'
                        else:
                            GICSvalue = str(sheet.cell(row=rowall, column=GICSID).value)
                            if str(GICSvalue).isdigit() != True:
                                GICSvalue = 'IS NULL'
                            else:
                                GICSvalue = '=' + str(sheet.cell(row=rowall, column=GICSID).value)
                        if (ProductID == 0):
                            Productvalue = 'IS NULL'
                        else:
                            Productvalue = str(sheet.cell(row=rowall, column=ProductID).value)
                            if str(Productvalue).isdigit() != True:
                                Productvalue = 'IS NULL'
                            else:
                                Productvalue = '=' + str(sheet.cell(row=rowall, column=ProductID).value)
                        if (subelementid == 0):
                            subelementvalue = 'IS NULL'
                        else:
                            subelementvalue = str(sheet.cell(row=rowall, column=subelementid).value)
                            if str(subelementvalue).isdigit() != True:
                                subelementvalue = 'IS NULL'
                            else:
                                subelementvalue = '=' + str(sheet.cell(row=rowall, column=subelementid).value)
                        if (othereleid == 0):
                            otherelevalue = 'IS NULL'
                        else:
                            otherelevalue = str(sheet.cell(row=rowall, column=othereleid).value)
                            if str(otherelevalue).isdigit() != True:
                                otherelevalue = 'IS NULL'
                            else:
                                otherelevalue = '=' + str(sheet.cell(row=rowall, column=othereleid).value)
                        # if (gxbrlvalue .isdigit() == True):
                        # print("gxbrlvalue",gxbrlvalue,"seg",seg,"geographyvalue",geographyvalue,"GICSvalue",GICSvalue,"Productvalue",Productvalue,"subelementvalue",subelementvalue,"otherelevalue",otherelevalue)
                        try:
                            if (gxbrlvalue.isdigit() == True):
                                # print(gxbrlvalue,seg)
                                try:
                                    checkifexist = "select * from DWElement where EleRefID= %s and SegmentID {}".format( \
                                        seg) + " and ProductID {}".format(Productvalue) + " and DWGeoID {}".format( \
                                        geographyvalue) + " and DWGroupID {}".format(
                                        GICSvalue) + " and OtherElementID {}".format( \
                                        otherelevalue) + " and SubElementID {}".format(subelementvalue) \
                                        # checkifexist = "select * from DWHMainQuery where GXbrlEleID = %s and SegmentID {}".format( \
                                    #         seg)

                                    cursor.execute(checkifexist, (gxbrlvalue,))
                                    dwhids = cursor.fetchall()
                                    coid = 3545
                                    print(dwhids)
                                    if dwhids:
                                        dwhid = str(dwhids[0][0])
                                        # if (dwhid == 12):
                                        periodconvert = 0
                                        periodconvert = periodcol + 1
                                        #
                                        for colperiod in range(periodconvert, max_col):
                                            value = (sheet.cell(row=rowall, column=colperiod).value)
                                            if value != None:
                                                # print(value)
                                                if int(value) != None:
                                                    # print("value", value)
                                                    periodvalue = (
                                                        sheet.cell(row=periodrow, column=colperiod).value)
                                                    periodvalue = periodvalue.replace(' ', '')

                                                    try:
                                                        vperiods = str(
                                                            repr(periodcolumn).replace('[', '(').replace(']',
                                                                                                         ')').replace(
                                                                "'", ''))

                                                        periodsearchquery = "SELECT PeriodID FROM CalendarPeriodNew WHERE %s in {}".format(
                                                            vperiods)
                                                        cursor.execute(periodsearchquery, (periodvalue,))
                                                        periodid = cursor.fetchall()
                                                        periodid = periodid[0][0]
                                                        periodid = 'P' + str(periodid)
                                                        print(dwhid, periodid, reporteddate, value)
                                                        searchQuery = "select * from DWPeriodSchemaExtract where DWID = %s and DWCoID = %s"%(dwhid,coid)
                                                        try:
                                                            cursor.execute(searchQuery)
                                                            re = cursor.fetchall()
                                                            print(re)
                                                            if re == []:
                                                                print(re)
                                                                insertQuery = "insert into DWPeriodSchemaExtract (DWID,ReportedDateID,DWCoID) values(%s,%s,%s)" % (
                                                                dwhid, reporteddate, coid,)
                                                                try:
                                                                    cursor.execute(insertQuery)
                                                                    connection.commit()
                                                                    print("query executed")
                                                                except Exception as e:
                                                                    print("error",e)
                                                        #
                                                            else:
                                                                print(re)
                                                                updateQuery = "Update DWPeriodSchemaExtract set `%s` = %s where DWID = %s and ReportedDateID = %s and DWCoID = %s "%(
                                                                        str(periodid), value, dwhid, reporteddate, coid)
                                                                try:
                                                                    cursor.execute(updateQuery)
                                                                    print("executed")
                                                                    connection.commit()
                                                                except Exception as e:
                                                                    print(e)


                                                        except Exception as e:
                                                            print("error",e)
                        #
                                                    except Exception as e:
                                                        print(e)
                        #
                                except Exception as e:
                                    print(e)
                        #
                        except Exception as e:
                            print(e)

            except Exception as e:
                print(e)
    except Exception as e:
        print(e)

# tablename = "`StarSchema1` (`FactID` smallint(6) NOT NULL, `UID` smallint(6) NOT NULL,  `PeriodID` varchar(50) NOT NULL,  `ReportedDateID` varchar(256) DEFAULT NULL,  `Value` int(11) DEFAULT NULL,  PRIMARY KEY (`FactID`)) ENGINE=InnoDB DEFAULT CHARSET=latin1;"
# dumpdataStarSchema()
# tablename = "`PeriodwiseSchema1` (`FactID` smallint(6) NOT NULL,`ReportedDateID` varchar(256) DEFAULT NULL,`DWHID` int (23), PRIMARY KEY (`FactID`)) ENGINE=InnoDB DEFAULT CHARSET=latin1;"
dumpPeriodSchema()
# # tablename = "`UIDwiseSchema1` (`FactID` smallint(6) NOT NULL,`PeriodID` varchar(50) NOT NULL, `ReportedDateID` varchar(256) DEFAULT NULL,  PRIMARY KEY (`FactID`)) ENGINE=InnoDB DEFAULT CHARSET=latin1;"
# dumpUIDSchema ()




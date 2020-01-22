import mysql.connector
from openpyxl import load_workbook
from datetime import date
try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='DWHDB',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)
#
#
        ticker = ['CDEV','ECA','EOG','JAG','LPI','PE','PXD']
        COID = [11569,6918,3545,11668,9020,9773,3741]
        for tic,co in zip(ticker,COID):
            wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW1StcokPrice.xlsx',data_only=True)      # print(result)
            sheet = wb.get_sheet_by_name(tic)
            max_col = (sheet.max_column)+1
            max_row = (sheet.max_row) + 1
            prevvalue= 0
            CloseValue  = 0    #
            try:
                for row in range(2, max_row):
                        val = (str(sheet.cell(row = row,column = 1).value)).strip("00:00:00").replace('-','')
                        yr = val[0:4]
                        mon = val[4:6]
                        date = val[6:8]
                        valid = val.replace('/','')
                        # print(valid)
                        try:
                            reporteddatequery = "select * from CalendarDate where replace(YYYYMMDD,'-','') = %s"
                            cursor.execute(reporteddatequery, (valid,))
                            reporteddate = cursor.fetchall()
                            reporteddate = reporteddate[0][0]
                            # print(reporteddate)
                        except Exception as e:
                            print(e)
                        dateddmmyy = str(date) + str(mon) + str(yr)
                        # print(dateddmmyy)
                        findQuery = "SELECT * FROM  CalendarDailyPeriod where REPLACE(PeriodName1, '/', '') = %s"%dateddmmyy
                        cursor.execute(findQuery)
                        res = cursor.fetchall()[0][0]
                        dateid = (res)
                        print(dateid,reporteddate)
                        Volume = (str(sheet.cell(row = row,column = 2).value))
                        CloseValue = (str(sheet.cell(row = row,column = 3).value))
                        LowValue = (str(sheet.cell(row = row,column = 4).value))
                        HighValue = (str(sheet.cell(row = row,column = 5).value))
                        coid = co
                        # print(Volume,CloseValue,LowValue,HighValue)
                        # # break
                        try:
                            #
                            insertQueryVolume = "INSERT INTO StocksStarSchemaNew(DWID,PeriodID,ReportedDateID,Value,DWCoID) VALUES (%s,%s,%s,%s,%s)" % (
                            '362', dateid, reporteddate, Volume, coid)
                            cursor.execute(insertQueryVolume)
                        #
                            insertQueryCloseValue = "INSERT INTO StocksStarSchemaNew(DWID,PeriodID,ReportedDateID,Value,DWCoID) VALUES (%s,%s,%s,%s,%s)" % (
                                '363', dateid, reporteddate, CloseValue, coid)
                            cursor.execute(insertQueryCloseValue)

                            insertQueryLowValue = "INSERT INTO StocksStarSchemaNew(DWID,PeriodID,ReportedDateID,Value,DWCoID) VALUES (%s,%s,%s,%s,%s)" % (
                                '364', dateid, reporteddate, LowValue, coid)
                            cursor.execute(insertQueryLowValue)

                            insertQueryHighValue = "INSERT INTO StocksStarSchemaNew(DWID,PeriodID,ReportedDateID,Value,DWCoID) VALUES (%s,%s,%s,%s,%s)" % (
                                '365', dateid, reporteddate, HighValue, coid)
                            cursor.execute(insertQueryHighValue)
                            connection.commit()
                        #
                        except Exception as e:
                            print(e)
                            pass



            except Exception as e:
                print(e)

except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")




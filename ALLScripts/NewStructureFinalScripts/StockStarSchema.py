import mysql.connector
from openpyxl import load_workbook
from datetime import date
try:
    connection = mysql.connector.connect(host='192.168.10.13',
                                         database='DWHDB',
                                         user='root',
                                         password='Bluecrest')
    if connection.is_connected():
        db_Info = connection.get_server_info()
        print("Connected to MySQL database... MySQL Server version on ", db_Info)
        cursor = connection.cursor(buffered=True)
        cursor.execute("select database();")
        record = cursor.fetchone()
        print("Your connected to - ", record)


        wb = load_workbook('/home/pavan/Music/DataWarehouse/DCExcelSheets/DW Excel/Stock Variability.xlsx',data_only=True)      # print(result)
        sheet = wb.get_sheet_by_name('EOG')
        max_col = (sheet.max_column)+1
        max_row = (sheet.max_row) + 1
        prevvalue= 0
        CloseValue  = 0
        try:
            today = date.today()
            reporteddatequery = "select * from CalendarDate where Date = %s"
            cursor.execute(reporteddatequery, (today,))
            reporteddate = cursor.fetchall()
            reporteddate = reporteddate[0][0]
            print(reporteddate)
        except Exception as e:
            print(e)
        try:
            for row in range(2, max_row):
                    val = (str(sheet.cell(row = row,column = 1).value)).strip("00:00:00").replace('-','')

                    findQuery = "SELECT * FROM  CalendarDate where REPLACE(Date, '-', '') = %s"%val
                    cursor.execute(findQuery)
                    res = cursor.fetchall()[0][0]
                    dateid = (res)
                    # Volume = (str(sheet.cell(row = row,column = 2).value))
                    # print(Volume)
                    # prevvalue = float(CloseValue)
                    # CloseValue = (str(sheet.cell(row = row,column = 3).value))
                    # LowValue = (str(sheet.cell(row = row,column = 4).value))
                    HighValue = (str(sheet.cell(row = row,column = 5).value))
                    # diff = float(CloseValue) - float(prevvalue)
                    dwhid = 283
                    coid = 32994
                    # print(dateid,diff)
                    # break
                    try:
                        #
                        # insertQuery = "INSERT INTO DWStocksDWHIDSchema(PeriodID,ReportedDateID,DWHCoID) VALUES (%s,%s,%s)" %( dateid, reporteddate,\
                        #                         coid)
                        # cursor.execute(insertQuery)
                        updateQuery = "Update DWStocksDWHIDSchema set `%s` = %s where PeriodID = %s and DWHCoID = %s"
                        cursor.execute(updateQuery,(dwhid,HighValue,dateid,coid,))
                        connection.commit()


                    except Exception as e:
                        print(e)
                        pass


        except Exception as e:
            print(e)

except Exception as e:
    print("Error while connecting to MySQL",e)

finally:
    # closing database connection.
    if (connection.is_connected()):
        cursor.close()
        connection.close()
        print("MySQL connection is closed")